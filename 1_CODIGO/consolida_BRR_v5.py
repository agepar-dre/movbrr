# -*- coding: utf-8 -*-
"""
Created on Wed Jan 31 10:40:36 2024

@author: cecil.skaleski, est.angelo

Consolida os arquivos de BRR confome disposto na Nota Técnica n° XX/2024.
"""

import pandas as pd
import numpy as np
import os
import glob
from natsort import os_sorted #conda install conda-forge::natsort
import datetime
import subprocess

import openpyxl
from openpyxl.styles import PatternFill
from copy import copy

import tkinter as tk
from tkinter import filedialog

np.set_printoptions(linewidth=np.inf)
pd.set_option('display.max_columns', 20)
pd.set_option('display.max_rows', 100)
pd.set_option('display.width', 100)
pd.set_option("display.precision", 6)
# Don't wrap repr(DataFrame) across additional lines
pd.set_option("display.expand_frame_repr", False)
pd.options.display.float_format = '{:.2f}'.format
pd.set_option("display.date_dayfirst", True)
pd.set_option('display.max_colwidth', 55)

#______________________________________________________________________________________
#Funções acessórias
def formats2(x):
    return ('R${:,.2f}'.format(x)).replace(",", "~").replace(".", ",").replace("~", ".")

def formats3(x):
    return "{:.2f}%".format(x*100)

def monta_path(abs_path, folder_path, fname):
    #Retorna o caminho absoluto para o arquivo
    return os.path.join(abs_path, folder_path+fname)

def lista_arquivos_dir(path_base, ext):
    #Lista arquivos com a extensão indicada em um diretório
    file_paths = []
    pasta_raiz = path_base
    formato = r'\*.' + ext
    lista_paths = os_sorted(glob.glob(os.path.expanduser(pasta_raiz + formato)))
    lista_files = [os.path.basename(x) for x in lista_paths]
    return lista_files

def escolhe_arq(titulo, filetypes, dir_ini):
    root = tk.Tk()
    root.wm_attributes('-topmost', 1)
    root.withdraw()
    file_path = filedialog.askopenfilename(parent=root, initialdir=dir_ini, title=titulo, filetypes=filetypes)
    #showinfo(title='Arquivo selecionado', message=filename))
    root.destroy()
    return file_path

def escolhe_pasta(titulo, dir_ini):
    root = tk.Tk()
    root.wm_attributes('-topmost', 1)
    root.withdraw()
    file_path = filedialog.askdirectory(parent=root, initialdir=dir_ini, title=titulo)
    root.destroy()
    return file_path
#______________________________________________________________________________________

#______________________________________________________________________________________
#Funções de processamento de dados
def tabela_resumo(df, contas, col_qtde, col_custo, col_conta, col_mun):
    #Calcula a tabela resumo
    qtdes = []
    muns = []
    custos = []
    itens = []
    for conta in contas:
        df_aux = df[df[col_conta] == conta]
        item = len(df_aux)
        if item > 0:
            qtde = df_aux[col_qtde].sum()
            nmun = len(df_aux[col_mun].unique())
            custo = df_aux[col_custo].sum()
        else:
            qtde = 0
            nmun = 0
            custo = 0
        #Registra
        qtdes.append(qtde)
        muns.append(nmun)
        custos.append(custo)
        itens.append(item)
    #Monta o dataframe
    df_resumo = pd.DataFrame({
        'Conta contábil': contas,
        'Linhas': itens,
        'Qtde de bens': qtdes,
        'N municípios': muns,
        'Custo contábil': custos,
        })
    #Ordena por maior custo, linhas, qtde e municipios
    df_resumo = df_resumo.sort_values(['Custo contábil', 'Linhas', 'Qtde de bens', 'N municípios'], ascending=[False, False, False, False]).reset_index(drop=True)
    #Calcula o impacto percentual e cumulativo
    df_resumo['%'] = df_resumo['Custo contábil'] / df_resumo['Custo contábil'].sum()
    df_resumo['% acum'] = df_resumo['%'].cumsum()
    return df_resumo


def iu(tipo, plaqueta, complemento):
    #Monta o código identificador único do ativo na base de referência da Sanepar
    return f"{str(tipo).split('.')[0]}-{str(plaqueta).split('.')[0]}-{str(complemento).split('.')[0]}"

def verifica_reqs(df, col_names):
    #Verifica a consistência das colunas indicadas (iu, taxa de depreciação e qtde)
    col_iu = col_names[0]
    col_deprec = col_names[1]
    col_qtde = col_names[2]
    flag = False
    err_msg = []
    #Verifica se a coluna existe
    if col_iu not in df.columns:
        err_msg.append('Coluna com identificador único (iu) não encontrada!')
        flag = True
    else:
        #Verifica se a coluna é consistente
        mask = (df[col_iu].isnull()) | (df[col_iu].apply(str).str.strip() == '')
        df_iu = df[mask]
        if len(df_iu) > 0:
            err_msg.append('Coluna identificador único (iu) inconsistente! (itens nulos ou vazios)')
            flag = True
        #Verifica se há itens com iu replicado
        df_aux = df.groupby(col_iu).size().reset_index(name='qtde')
        df_rep = df_aux[df_aux['qtde'] > 1]
        if len(df_rep) > 0:
            err_msg.append('Identificadores únicos (iu) replicados nos dados processados!')
            flag = True 
        #Verifica a composição do código identificador único de ativo (IU)
        df_comp = df.copy()
        df_comp['iu_verif'] = df_comp.apply(lambda x: iu(x['tipo'], x['plaqueta'], x['complemento']), axis=1)
        #Compara
        df_comp_res = df_comp[df_comp['iu_verif'] != df_comp[col_iu]]
        if len(df_comp_res) > 0:
            #Itens com código inadequado
            print('')
            print('Inconsistências detectadas na composição do identificador único: ')
            print(df_comp_res.loc[:, ['tipo', 'plaqueta', 'complemento', col_iu, 'iu_verif']])
            err_msg.append('Inconsistências detectadas na composição do identificador único!')
            flag = True
    #Itens sem taxa de depreciação
    #Verifica se a coluna existe
    if col_deprec not in df.columns:
        err_msg.append('Coluna taxa de depreciação não encontrada!')
        flag = True
    else:
        #Verifica se a coluna é consistente
        mask = (df[col_deprec].isnull()) | (df[col_deprec].apply(str).str.strip() == '')
        df_iu = df[mask]
        if len(df_iu) > 0:
            err_msg.append('Coluna taxa de depreciação inconsistente! (itens nulos ou vazios)')
            flag = True
    #Itens sem quantidade definida
    #Verifica se a coluna existe
    if col_qtde not in df.columns:
        err_msg.append('Coluna quantidade não encontrada!')
        flag = True
    else:
        #Verifica se a coluna é consistente
        mask = (df[col_qtde].isnull()) | (df[col_qtde].apply(str).str.strip() == '')
        df_iu = df[mask]
        if len(df_iu) > 0:
            err_msg.append('Coluna quantidade inconsistente! (itens nulos ou vazios)')
            flag = True
    return flag, err_msg
#______________________________________________________________________________________

#______________________________________________________________________________________
#Replica formatação do arquivo excel de saída
def num_para_letra(n):
    letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if n <= 26:
        result = letras[n - 1]
    else:
        quociente = (n - 1) // 26
        resto = (n - 1) % 26
        result =  num_para_letra(quociente) + letras[resto]
    return result

def justificar_tamanho_colunas(ws):
    for coluna in ws.columns:
        max_length = 0
        for cell in coluna[1:]:
            if cell.value:
                if isinstance(cell.value, float):
                    if 0 < cell.value < 1:
                        max_length = max(max_length, 7)
                    else:
                        max_length = max(max_length, len(str(cell.value)))
                else:
                    max_length = max(max_length, len(str(cell.value)))

        largura_justificada = (max_length + 2) * 1.2
        ws.column_dimensions[coluna[0].column_letter].width = largura_justificada
    return

def copiar_estilos_celula(origem, destino):
    destino.font = copy(origem.font)
    destino.fill = copy(origem.fill)
    destino.border = copy(origem.border)
    destino.alignment = copy(origem.alignment)
    destino.protection = copy(origem.protection)
    destino.number_format = origem.number_format
    return

def copia_format(template_path, original_path, output_path):
    template_workbook = openpyxl.load_workbook(template_path)
    original_workbook = openpyxl.load_workbook(original_path)

    template_sheet = template_workbook.active
    original_sheet = original_workbook.active
    
    tpdebug = round(len(list(original_sheet.iter_rows()))/2)
    template_place = list(template_sheet.iter_rows())[1:] * round((len(list(original_sheet.iter_rows()))-1)/round(len(list(template_sheet.iter_rows())[1:])))
    new_template_list = [list(template_sheet.iter_rows())[0]] * round(len(list(original_sheet.iter_rows())[0])/len(list(original_sheet.iter_rows())[0]))
    new_template_list.extend(template_place)

    justificar_tamanho_colunas(original_sheet)

    lenght_x = len(new_template_list[0])
    lenght_y = len(new_template_list)

    for sheet_name in original_workbook.sheetnames:
        ws = original_workbook[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=lenght_y*3, min_col=1, max_col=lenght_x+100):
            for cell in row:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for row_template, row_original in zip(new_template_list, original_sheet.iter_rows()):
        for cell_template, cell_original in zip(row_template, row_original):
            copiar_estilos_celula(cell_template, cell_original)

    original_workbook.save(output_path)
    return
#______________________________________________________________________________________

#______________________________________________________________________________________
#Função principal
def consolida_BRR(export, open_folder, abs_path, path_base, mensagem_text):
    """Consolida os arquivos de BRR parcial conforme orientação técnica (protocolo n° xx.xxx.xx-x)"""
    #Habilita o feedback na caixa de mensagens
    mensagem_text.config(state=tk.NORMAL)
    #_______________________________________
    #Carrega as bases de dados n pasta indicada
    print('')
    print('_____________________________________CONSOLIDA BRR_____________________________________')
    print('Carregando as bases de dados...')
    mensagem_text.insert(tk.END, f'Carregando as bases de dados...\n')
    mensagem_text.see('end')
    mensagem_text.update()
    #Lista todos os arquivos da pasta
    lista_arqs = lista_arquivos_dir(path_base, '*')
    lista_dfs = []
    col_names = []
    col_fname = []
    df_ius = pd.DataFrame([])
    flag = False
    for fname in lista_arqs:
        path_ref = os.path.join(path_base, fname)
        print(f'    Arquivo: {fname}')
        mensagem_text.insert(tk.END, f'    Arquivo: {fname}\n')
        aux_df = pd.read_excel(path_ref)
        #Faz o drop de eventuais valores expurios
        aux_df.dropna(how='all', axis=1, inplace=True)
        aux_df.dropna(how='all', axis=0, inplace=True)
        #Verifica se há coluna com IU
        if 'iu' not in aux_df.columns:
            print(f'        Coluna com identificador único (iu) não encontrada!')
            mensagem_text.insert(tk.END, f'        Coluna com identificador único (iu) não encontrada!\n')
            mensagem_text.see('end')
            mensagem_text.update()
            flag = True
        else:
            #Registra
            df_ius = pd.concat([df_ius, aux_df['iu']], axis=0, ignore_index=True)
            #Conta a frequencia dos identificadores únicos
            df_ius_gb = df_ius.groupby(0).size().reset_index(name='qtde')
            #Verifica se há itens com iu replicado
            mask = (df_ius_gb['qtde'] > 1)
            df_rep = df_ius_gb[mask]
            nrep = len(df_rep)
            if nrep > 0:
                print(f'        {nrep} identificadores únicos (iu) replicados nos dados processados!')
                mensagem_text.insert(tk.END, f'        {nrep} identificadores únicos (iu) replicados nos dados processados!\n')
                mensagem_text.see('end')
                mensagem_text.update()
                flag = True
        #Registra
        lista_dfs.append(aux_df)
        cols = aux_df.columns.to_list()
        col_names += cols
        col_fname += [fname]*len(cols)
        print(f'    Colunas: ')
        mensagem_text.insert(tk.END, '    Colunas: \n')
        mensagem_text.see('end')
        mensagem_text.update()
        for col in cols:
            print(f'        {col}')
            mensagem_text.insert(tk.END, f'        {col}\n')
            mensagem_text.see('end')
            mensagem_text.update()
    #_______________________________________
    
    #_______________________________________
    #Somente continua se houver consistência mínima de dados
    if flag == False:
        #_______________________________________
        #Verifica se o nome das colunas são idênticos
        #Busca os nomes comuns a todos os dataframes
        df_col_names = pd.DataFrame({
            'nome_coluna': col_names,
            'arquivo': col_fname,
            })
        #Agrupa
        df_col_gp = df_col_names.groupby('nome_coluna').size().reset_index(name='n_arquivos')
        narqs = len(lista_arqs)
        #Classifica as colunas (colunas comuns e não comuns)
        df_col_gp['tipo'] = 'Não comum'
        mask = df_col_gp['n_arquivos'] == narqs
        df_col_gp.loc[mask, 'tipo'] = 'Comum'
        #Filtra as colunas não comuns
        df_col_nc = df_col_gp[df_col_gp['tipo'] == 'Não comum'].copy()
        flag_cont = 'N'
        if len(df_col_nc) > 0:
            #Adiciona informações de nomes de arquivos
            df_col_nc['arquivo'] = 'Vários'
            mask =( df_col_nc['n_arquivos'] == 1)
            df_col_uniqs = df_col_nc.loc[mask, 'nome_coluna']
            if len(df_col_uniqs) > 0:
                #Nomes de colunas que só aparecem em um único arquivo
                col_uniqs = df_col_uniqs.to_list()
                #Concatena as informações do dataframe de colunas e arquivos
                #Dataframe de origem
                df_aux1 = df_col_names[df_col_names['nome_coluna'].isin(col_uniqs)]
                df_aux1.index = df_aux1['nome_coluna']
                #Dataframe de destino
                df_col_nc.loc[mask, 'arquivo'] = df_aux1.loc[col_uniqs, 'arquivo'].to_list()
            print(' ')
            print(f'{len(df_col_nc)} colunas não comuns a todos os arquivos encontradas: ')
            print(df_col_nc)
            print(' ')
            mensagem_text.insert(tk.END, '\n')
            mensagem_text.see('end')
            mensagem_text.insert(tk.END, f'{len(df_col_nc)} colunas não comuns a todos os arquivos encontradas: \n')
            mensagem_text.see('end')
            mensagem_text.insert(tk.END, f'{df_col_nc}\n')
            mensagem_text.see('end')
            mensagem_text.update()
            flag_cont = (input('Continuar mesmo assim? (S ou N) ')).upper().strip()
        else:
            flag_cont = 'S'
        #_______________________________________
        
        #_______________________________________
        #Verifica se continua a execução
        if flag_cont == 'S':
            #_______________________________________
            #Concatena os dados
            print('')
            print('Concatenando as bases...')
            print('Verificando requisitos mínimos...')
            mensagem_text.insert(tk.END, '\n')
            mensagem_text.see('end')
            mensagem_text.insert(tk.END, 'Concatenando as bases...\n')
            mensagem_text.see('end')
            mensagem_text.insert(tk.END, 'Verificando requisitos mínimos...\n')
            mensagem_text.see('end')
            mensagem_text.update()
            df_brr = pd.DataFrame([])
            idx = 0
            flag = False
            flag_cont2 = 'S'
            for df in lista_dfs:
                #_______________________________________
                #Verificação de requisitos mínimos
                print(f'    Arquivo {lista_arqs[idx]}')
                mensagem_text.insert(tk.END, f'    Arquivo {lista_arqs[idx]}\n')
                mensagem_text.see('end')
                mensagem_text.update()
                flag, err_msg = verifica_reqs(df, ['iu', 'taxa_deprec_anos', 'qtde'])
                #Verifica se continua a execução
                if flag == True:
                    for msgm in err_msg:
                        print(msgm)
                    print('')
                    flag_cont2 = (input('Requisitos de consistência não cumpridos. Continuar mesmo assim? (S ou N) ')).upper().strip()
                    #flag_cont2 = (mensagem_text.wait_variable('Requisitos de consistência não cumpridos. Continuar mesmo assim? (S ou N) ').get()).upper().strip()
                else:
                    print('        ok!')
                #Se a decisão for prosseguir
                if flag_cont2 == 'S':
                    #Concatena
                    df_brr = pd.concat([df_brr, df], ignore_index=True, axis=0)
                    #Registra
                    idx += 1
                else:
                    break
                #_______________________________________
            #_______________________________________
            
            #_______________________________________
            #Verifica se continua a execução
            if flag_cont2 == 'S':
                #_______________________________________
                #Cria tabela resumo por conta contábil
                #Avalia a quantidade de itens e os valores totais por conta contábil
                col_qtde = 'qtde'
                col_custo = 'custo_contabil'
                col_conta = 'conta_contabil'
                col_mun = 'municipio'
                
                #Monta a tabela de resumo
                contas = df_brr[col_conta].unique()
                df_res_ref = tabela_resumo(df_brr, contas, col_qtde, col_custo, col_conta, col_mun)
                #Adiciona os totais
                df_res_ref.loc['TOTAL', :] = df_res_ref.sum(axis=0)
                df_res_ref.loc['TOTAL', 'Conta contábil'] = 'TOTAL'
                df_res_ref.loc['TOTAL', 'N municípios'] = len(df_brr[col_mun].unique())
                df_res_ref.loc['TOTAL', '% acum'] = 1
                #Formata valores unitários
                df_res_ref['Linhas'] = df_res_ref['Linhas'].apply(int)
                df_res_ref['Qtde de bens'] = df_res_ref['Qtde de bens'].apply(int)
                df_res_ref['N municípios'] = df_res_ref['N municípios'].apply(int)
                
                #Apresenta o resultado
                print('')
                print('Tabela resumo da base de referência')
                mensagem_text.insert(tk.END, '\n')
                mensagem_text.see('end')
                mensagem_text.insert(tk.END, 'Tabela resumo da base de referência\n')
                mensagem_text.see('end')
                df_res_ref_f = df_res_ref.copy()
                #Formata valores monetários
                df_res_ref_f['Custo contábil'] = df_res_ref_f['Custo contábil'].apply(formats2)
                #Formata valores percentuais
                df_res_ref_f['%'] = df_res_ref_f['%'].apply(formats3)
                df_res_ref_f['% acum'] = df_res_ref_f['% acum'].apply(formats3)
                print(df_res_ref_f)
                mensagem_text.insert(tk.END, f'{df_res_ref_f}\n')
                mensagem_text.see('end')
                mensagem_text.update()
                #_______________________________________
                
                #_______________________________________
                #Exporta resultados
                n_linhas = len(df_brr)
                data_hj = datetime.datetime.today().strftime('%d-%m-%Y_%Hh%Mmin%Ss')
                rtp = df_brr['rtp'].apply(int).max()
                #Exporta o resumo da BRR em formato de planilha excel
                export = True
                if export == True:
                    print('')
                    print(f'Exportando dados em formato .xlsx...')
                    mensagem_text.insert(tk.END, '\n')
                    mensagem_text.insert(tk.END, f'Exportando dados em formato .xlsx...\n')
                    mensagem_text.see('end')
                    mensagem_text.update()
                    fname = f"RESUMO_BRR_{rtp}RTP_{str(n_linhas)}_itens_{data_hj}.xlsx"
                    folder_path = '4_SAIDA_CONSOLIDA//'
                    folder_exp = folder_path.split('/')[0]
                    path_exp = monta_path(abs_path, folder_path, fname)
                    with pd.ExcelWriter(path_exp) as writer:
                        df_res_ref.to_excel(writer, index=False)
                    print(f'    Arquivo exportado com sucesso!')
                    print(f'    {path_exp}')
                    mensagem_text.insert(tk.END, f'    Arquivo exportado com sucesso!\n')
                    mensagem_text.insert(tk.END, f'    {path_exp}\n')
                    mensagem_text.see('end')
                    mensagem_text.update()
                    #Replica a formatação do template
                    print('Replicando a formatação do arquivo de modelo...')
                    mensagem_text.insert(tk.END, "Replicando a formatação do arquivo de modelo...\n")
                    mensagem_text.see('end')
                    mensagem_text.update()
                    folder_path = '3_ENTRADA_CONSOLIDA/1_FORMATOS//'
                    fname = 'Template_resumo_brr.xlsx'
                    template_path = monta_path(abs_path, folder_path, fname)
                    copia_format(template_path, path_exp, path_exp)
                    print('    Formatação replicada com sucesso!')
                    mensagem_text.insert(tk.END, "    Formatação replicada com sucesso!\n")
                    mensagem_text.see('end')
                    mensagem_text.update()
                    
                #Exporta a BRR em formato excel
                    print('')
                    print(f'Exportando dados em formato .xlsx...')
                    mensagem_text.insert(tk.END, '\n')
                    mensagem_text.insert(tk.END, f'Exportando dados em formato .xlsx...\n')
                    mensagem_text.see('end')
                    mensagem_text.update()
                    fname = f"BRR_{rtp}RTP_{str(n_linhas)}_itens_{data_hj}.xlsx"
                    folder_path = '4_SAIDA_CONSOLIDA//'
                    path_exp = monta_path(abs_path, folder_path, fname)
                    with pd.ExcelWriter(path_exp) as writer:
                        df_brr.to_excel(writer, index=False)
                    print(f'    Arquivo exportado com sucesso!')
                    print(f'    {path_exp}')
                    mensagem_text.insert(tk.END, f'    Arquivo exportado com sucesso!\n')
                    mensagem_text.insert(tk.END, f'    {path_exp}\n')
                    mensagem_text.see('end')
                    mensagem_text.update()
                #Abre a pasta com os arquivos gerados
                    if open_folder == True:
                        res = subprocess.Popen(fr'explorer "{folder_exp}"')
                #_______________________________________
        #_______________________________________
    else:
        print('')
        print('Dados inconsistentes! Verificar os arquivos de entrada!')
        mensagem_text.insert(tk.END, '\n')
        mensagem_text.insert(tk.END, 'Dados inconsistentes! Verificar os arquivos de entrada!\n')
        mensagem_text.see('end')
        mensagem_text.update()
    print('_____________________________________CONSOLIDA BRR_____________________________________')
    #_______________________________________
    #Desabilita o feedback no campo de mensagens
    mensagem_text.config(state=tk.DISABLED) 
    return
#______________________________________________________________________________________

#______________________________________________________________________________________
#Funções GUI
def buscar_pasta(entry_pasta, path_var):
    abs_path = os.path.dirname(__file__)
    folder_path = '3_ENTRADA_CONSOLIDA'
    dir_ini = os.path.join(abs_path, folder_path)
    pasta = filedialog.askdirectory(initialdir=dir_ini, title='Selecione a pasta com os arquivos de BRR a consolidar')
    entry_pasta.delete(0, tk.END)
    entry_pasta.insert(0, pasta)
    entry_pasta.xview_moveto(1)
    path_var.set(pasta)

def consolidar_brr(path_base_var, mensagem_text):
    #Captura o conteúdo dos elementos de texto
    path_base = path_base_var.get()
    #______________________________________EXECUÇÃO________________________________________
    #Seleciona os arquivos de entrada
    abs_path = os.path.dirname(__file__)
    export = True
    open_folder = True
    consolida_BRR(export, open_folder, abs_path, path_base, mensagem_text)
    #______________________________________EXECUÇÃO________________________________________
    #Feedback na caixa de mensagens
    mensagem_text.config(state=tk.NORMAL)
    mensagem_text.insert(tk.END, "BRR consolidada com sucesso!\n")
    mensagem_text.config(state=tk.DISABLED)
    
def make_frame(frame):
    #Captura o conteudo da caixa de texto com o caminho da pasta selecionada
    path_base_var = tk.StringVar()

    # Selecionar Pasta de arquivos
    label_arquivo = tk.Label(frame, text="Pasta de arquivos:", fg="blue")
    label_arquivo.grid(row=0, column=0, padx=10, pady=(10,0), sticky="w")

    entry_pasta = tk.Entry(frame, width=70, bg='lightgrey')
    entry_pasta.grid(row=0, column=1, pady=(10,0), sticky="w")

    btn_buscar = tk.Button(frame, text="Buscar", bg="navy", fg="white", command=lambda: buscar_pasta(entry_pasta, path_base_var))
    btn_buscar.grid(row=0, column=2, padx=5, pady=(10,0), sticky="w")

    # Botão Converter Base de Dados
    btn_converter_base_dados = tk.Button(frame, text="Consolidar BRR", bg="navy", fg="white", width=20, height=2, command=lambda: consolidar_brr(path_base_var, mensagem_text))
    btn_converter_base_dados.grid(row=1, column=1, columnspan=3, pady=10, sticky="e")

    # Display
    mensagem_label = tk.Label(frame, text="Mensagens:", fg="blue")
    mensagem_label.grid(row=6, column=0, pady=(132, 0), padx=10, sticky="w")  

    mensagem_text = tk.Text(frame, height=3, width=80, bg="white", bd=1, relief="solid")
    mensagem_text.grid(row=7, column=0, columnspan=3, padx=10, pady=5, sticky="ew")
    mensagem_text.config(state=tk.DISABLED)

    scrollbar = tk.Scrollbar(frame, command=mensagem_text.yview)
    scrollbar.grid(row=7, column=3, rowspan=3, sticky='nse')
    mensagem_text['yscrollcommand'] = scrollbar.set

def init_frame():
    root = tk.Tk()
    root.title("Consolida BRR")

    largura_janela = 700
    altura_janela = 220

    largura_tela = root.winfo_screenwidth()
    altura_tela = root.winfo_screenheight()

    posx = largura_tela // 2 - largura_janela // 2
    posy = altura_tela // 2 - altura_janela // 2

    root.geometry(f"{largura_janela}x{altura_janela}+{posx}+{posy}")

    frame = tk.Frame(root)
    frame.pack(pady=20)

    make_frame(frame)

    root.mainloop()
