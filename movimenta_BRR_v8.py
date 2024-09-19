# -*- coding: utf-8 -*-
"""
Created on Wed Jan 31 10:40:36 2024

@author: cecil.skaleski, est.angelo

Movimenta a BRR confome disposto na Nota Técnica n° XX/2024.
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
from dateutil.relativedelta import relativedelta
import datetime
import matplotlib.dates as md
import time
import calendar
import math
import subprocess

import openpyxl
from openpyxl.styles import PatternFill
from copy import copy

import tkinter as tk
from tkinter import filedialog

import numpy_financial as npf
    #conda install -c conda-forge numpy-financial

np.set_printoptions(linewidth=np.inf)
pd.set_option('display.max_columns', 20)
pd.set_option('display.max_rows', 100)
pd.set_option('display.width', 100)
pd.set_option("display.precision", 6)
# Don't wrap repr(DataFrame) across additional lines
pd.set_option("display.expand_frame_repr", False)
pd.options.display.float_format = '{:.2f}'.format
pd.set_option("display.date_dayfirst", True)
pd.set_option('display.max_colwidth', 55)

#______________________________________________________________________________________
#Funções acessórias
def formats2(x):
    return ('R${:,.2f}'.format(x)).replace(",", "~").replace(".", ",").replace("~", ".")

def formats3(x):
    return "{:.2f}%".format(x*100)

def millify_rs(n):
    millnames = ['',' Mil',' Milhões',' Bilhões',' Trilhões']
    n = float(n)
    millidx = max(0,min(len(millnames)-1,
                        int(math.floor(0 if n == 0 else math.log10(abs(n))/3))))
    return 'R${:.2f}{}'.format(n / 10**(3 * millidx), millnames[millidx]).replace('.', ',')
 
def monta_path(abs_path, folder_path, fname):
    #Retorna o caminho absoluto para o arquivo
    return os.path.join(abs_path, folder_path+fname)

def escolhe_arq(titulo, filetypes, dir_ini):
    root = tk.Tk()
    root.wm_attributes('-topmost', 1)
    root.withdraw()
    file_path = filedialog.askopenfilename(parent=root, initialdir=dir_ini, title=titulo, filetypes=filetypes)
    #showinfo(title='Arquivo selecionado', message=filename))
    root.destroy()
    return file_path

def importar_arquivo(file_path):
    if file_path.endswith('.xlsx'):
        return pd.read_excel(file_path)
    elif file_path.endswith('.csv'):
        return pd.read_csv(file_path)
    elif file_path.endswith('.hdf'):
        return pd.read_hdf(file_path)
    elif file_path.endswith('.feather'):
        return pd.read_feather(file_path)
    else:
        raise ValueError("Formato de arquivo não suportado: Use .xlsx, .csv, .hdf, ou .feather")

def exportar_arquivo(df, output_path, output_format):
    if output_format == "xlsx":
        with pd.ExcelWriter(output_path) as writer:
            df.to_excel(writer, index=False)
    elif output_format == "csv":
        df.to_csv(output_path, index=False)
    elif output_format == "hdf":
        df.to_hdf(output_path, key='df', mode='w')
    elif output_format == "feather":
        df.to_feather(output_path)
    else:
        raise ValueError(f"Formato de saída '{output_format}' não suportado!")
#______________________________________________________________________________________

#______________________________________________________________________________________
#Funções de processamento de dados
def agrupa2(database, i_colunas_agrupamento, i_coluna_agregacao, ordem_decrescente, formatar):
    #Agrupa os dados por uma coluna específica e retorna as colunas selecionadas

    tam = len(i_colunas_agrupamento)
    #Constroi o filtro de colunas de agrupamento dinamicamente
    colunas_agrupamento = '['
    for i in i_colunas_agrupamento:
        if len(i_colunas_agrupamento) > 1 and i != i_colunas_agrupamento[len(i_colunas_agrupamento)-1]:
            colunas_agrupamento = colunas_agrupamento + "database.columns[" + str(i) + '], '
        else:
            colunas_agrupamento = colunas_agrupamento + "database.columns[" + str(i) + ']'
    colunas_agrupamento = colunas_agrupamento + ']'

    #Constroi o comando para agrupamento:
    comando_agrup = "database.groupby(" + colunas_agrupamento + ")[database.columns[i_coluna_agregacao]].sum().reset_index(name=database.columns[i_coluna_agregacao])"
    df_agrupado = eval(comando_agrup)

    if ordem_decrescente == 1:
        df_agrupado = df_agrupado.sort_values(database.columns[i_coluna_agregacao], ascending=False).reset_index(drop=True)

    #Calcula o percentual
    df_agrupado.loc[:,'%'] = df_agrupado.iloc[:,tam]/df_agrupado.iloc[:,tam].sum()
    #Calcula o acumulado
    df_agrupado.loc[:,'Acumulado'] = df_agrupado.iloc[:,tam+1].cumsum()

    if formatar == 1:
        #Formata como dinheiro
        df_agrupado.iloc[:,tam] = df_agrupado.iloc[:,tam].apply(formats2)
        #Formata os percentuais
        df_agrupado.iloc[:,tam+1] = df_agrupado.iloc[:,tam+1].apply(formats3)
        df_agrupado.iloc[:,tam+2] = df_agrupado.iloc[:,tam+2].apply(formats3)
    return df_agrupado
  
def calc_brr_imob2(df_base, db_mov, df_eleg):
    #Calcula os resultados da BRR para uma determinada database de depreciação dos ativos (considerando alterações de elegibilidade)
    
    #Data do início do exercício
    data_ini = pd.to_datetime(f'01/01/{db_mov.year}')
    #Filtra os ativos imobilizados até a data especificada
    df_aux = df_base[df_base['data_imob'] <= db_mov].copy()
    #Indexa pelo identificador único
    df_aux.index = df_aux['iu']
    
    #Atualiza elegibilidade
    #Verifica se há itens imobilizados com atualização de elegibilidade no período
    df_aux_eleg = df_eleg[df_eleg['data'] < db_mov].copy()
    #Verifica se os itens já foram imobilizados na BRR
    df_aux_eleg = df_aux_eleg[df_aux_eleg['iu'].isin(df_aux['iu'].to_list())]
    if len(df_aux_eleg) > 0:
        mask = (df_aux_eleg['elegibilidade'] == 'Não elegível')
        df_aux_eleg = df_aux_eleg[mask].copy()
        df_aux_eleg.index = df_aux_eleg['iu']
        
    #Calcula a vida útil do ativo [anos]
    df_aux['vur_anos'] = 1 / (df_aux['taxa_deprec_anos'] / 100)
    df_aux['tdr_percent_ano'] = (df_aux['taxa_deprec_anos'] / 100)
    df_aux['tdr_monet_ano'] = (df_aux['taxa_deprec_anos'] / 100) * df_aux['vrb']

    #Database da base (depreciação)
    #Ativos elegíveis
    df_aux['data_mov_atual'] = db_mov
    
    #Restringe a data de movimentação de ativos não elegíveis
    if len(df_aux_eleg) > 0:
        mask = df_aux_eleg.index
        df_aux.loc[mask, 'data_mov_atual'] = df_aux_eleg['data'].to_list()
        df_aux.loc[mask, 'elegibilidade'] = df_aux_eleg.loc[mask, 'elegibilidade']
        
    #Calcula vida útil consumida
    df_aux['vur_consumida_anos'] = (df_aux['data_mov_atual'] - df_aux['data_imob']) / pd.Timedelta('365 days') #Ano fiscal considerado: ano civil
    #Vida útil consumida não pode ser negativa
    mask = (df_aux['vur_consumida_anos'] < 0)
    df_aux.loc[mask, 'vur_consumida_anos'] = 0
    #Vida útil consumida não pode ser maior que a vida útil do ativo
    mask = (df_aux['vur_consumida_anos'] > df_aux['vur_anos'])
    df_aux.loc[mask, 'vur_consumida_anos'] = df_aux.loc[mask, 'vur_anos']
    
    #Depreciação acumulada
    df_aux['dep_reg_acum'] = df_aux['vur_consumida_anos'] * df_aux['tdr_monet_ano']
    #Calcula a o valor regulatório líquido dos ativos
    df_aux['vrl'] = df_aux['vrb'] - df_aux['dep_reg_acum']

    #_______________________________________
    #Calculo da BRR
    #Ignora ativos 100% amortizados e não elegíveis no exercício considerado
    mask = (df_aux['vur_consumida_anos'] < df_aux['vur_anos']) & (df_aux['elegibilidade'] != 'Não elegível')
    brr_bruta = df_aux.loc[mask, 'vrb'].sum()
    brr_liquida = df_aux.loc[mask, 'vrl'].sum()
    dep_acum = df_aux['dep_reg_acum'].sum()
    
    #Calcula os investimentos imobilizados no período
    mask = (df_aux['data_imob'] >= data_ini) & (df_aux['data_imob'] <= db_mov)
    invest = df_aux.loc[mask, 'vrb'].sum()
    
    #Calcula o saldo inelegível no período
    mask = ((df_aux['vur_consumida_anos'] < df_aux['vur_anos']) & (df_aux['elegibilidade'] == 'Não elegível')) & ((df_aux['data_mov_atual'] >= data_ini) & (df_aux['data_mov_atual'] <= db_mov))
    brr_bruta_ine = df_aux.loc[mask, 'vrb'].sum()
    brr_liquida_ine = df_aux.loc[mask, 'vrl'].sum()
    
    #Registra
    db_monet = pd.to_datetime(df_aux['data_monet_atual'].unique()[0])
    result_ano = [db_mov.strftime('%d/%m/%Y'), db_monet.strftime('%d/%m/%Y'), invest, brr_bruta, brr_liquida, dep_acum, brr_bruta_ine, brr_liquida_ine]
    #_______________________________________
    return df_aux, result_ano

def calc_brr_imob(df_base, db_mov):
    #Calcula os resultados da BRR para uma determinada database de depreciação dos ativos

    #Data do início do exercício
    data_ini = pd.to_datetime(f'01/01/{db_mov.year}')
    #Filtra os ativos imobilizados até a data especificada
    df_aux = df_base[df_base['data_imob'] <= db_mov].copy()
    #Indexa pelo identificador único
    df_aux.index = df_aux['iu']
        
    #Calcula a vida útil do ativo [anos]
    df_aux['vur_anos'] = 1 / (df_aux['taxa_deprec_anos'] / 100)
    df_aux['tdr_percent_ano'] = (df_aux['taxa_deprec_anos'] / 100)
    df_aux['tdr_monet_ano'] = (df_aux['taxa_deprec_anos'] / 100) * df_aux['vrb']

    #Database da base (depreciação)
    #Ativos elegíveis
    df_aux['data_mov_atual'] = db_mov
        
    #Calcula vida útil consumida
    df_aux['vur_consumida_anos'] = (df_aux['data_mov_atual'] - df_aux['data_imob']) / pd.Timedelta('365 days') #Ano fiscal considerado: ano civil
    #Vida útil consumida não pode ser negativa
    mask = (df_aux['vur_consumida_anos'] < 0)
    df_aux.loc[mask, 'vur_consumida_anos'] = 0
    #Vida útil consumida não pode ser maior que a vida útil do ativo
    mask = (df_aux['vur_consumida_anos'] > df_aux['vur_anos'])
    df_aux.loc[mask, 'vur_consumida_anos'] = df_aux.loc[mask, 'vur_anos']
    
    #Depreciação acumulada
    df_aux['dep_reg_acum'] = df_aux['vur_consumida_anos'] * df_aux['tdr_monet_ano']
    #Calcula a o valor regulatório líquido dos ativos
    df_aux['vrl'] = df_aux['vrb'] - df_aux['dep_reg_acum']

    #_______________________________________
    #Calculo da BRR
    #Ignora ativos 100% amortizados e não elegíveis
    mask = (df_aux['vur_consumida_anos'] < df_aux['vur_anos']) & (df_aux['elegibilidade'] != 'Não elegível')
    brr_bruta = df_aux.loc[mask, 'vrb'].sum()
    brr_liquida = df_aux.loc[mask, 'vrl'].sum()
    dep_acum = df_aux['dep_reg_acum'].sum()
    
    #Calcula os investimentos imobilizados no período
    mask = (df_aux['data_imob'] >= data_ini) & (df_aux['data_imob'] <= db_mov)
    invest = df_aux.loc[mask, 'vrb'].sum()
    
    #Calcula o saldo inelegível no período
    mask = ((df_aux['vur_consumida_anos'] < df_aux['vur_anos']) & (df_aux['elegibilidade'] == 'Não elegível')) & ((df_aux['data_mov_atual'] >= data_ini) & (df_aux['data_mov_atual'] <= db_mov))
    brr_bruta_ine = df_aux.loc[mask, 'vrb'].sum()
    brr_liquida_ine = df_aux.loc[mask, 'vrl'].sum()
    
    #Registra
    db_monet = pd.to_datetime(df_aux['data_monet_atual'].unique()[0])
    result_ano = [db_mov.strftime('%d/%m/%Y'), db_monet.strftime('%d/%m/%Y'), invest, brr_bruta, brr_liquida, dep_acum, brr_bruta_ine, brr_liquida_ine]
    #_______________________________________
    return df_aux, result_ano

def valida_data(data):
    #Função para testar se a data está no formato adequado
    try:
        if data != datetime.strptime(data, "%d/%m/%Y").strftime('%d/%m/%Y'):
            raise ValueError
            return True
    except ValueError:
        return False
    
def monta_data(mes, ano):
    #Monta uma data considerando o ano e o ultimo da do mês indicados
    last_day = calendar.monthrange(int(ano), int(mes))[1]
    return f'{last_day}/{mes}/{ano}'

def gera_datas(data_ini, data_fim):
    #Gera as datas intermediarias entre duas datas
    datas = [data_ini]
    delta_dias = (data_fim - data_ini).days
    for i in range(0, delta_dias):
        datas.append(datas[-1] + pd.Timedelta('1 day'))
    return datas        

def importa_ipca(path_ipca):
    #Importa o arquivo excel do IBGE com as séries históricas
    #https://www.ibge.gov.br/estatisticas/economicas/precos-e-custos/9256-indice-nacional-de-precos-ao-consumidor-amplo.html?=&t=series-historicas

    #Importa arquivo excel (alterar a engine e comando de importação para outros formatos)
    dfs = pd.read_excel(path_ipca, sheet_name=0)
    #path = 'F:\HOMEOFFICE\SANEAMENTO\\ipca_202011SerieHist.xls'
    #Seleciona as 3 primeiras colunas, remove as linhas com valores NaN e também as duas ulltimas linhas de texto do dataframe
    dfs = dfs.iloc[:, 0:3].dropna(how='all')
    #Cria uma mascara marcando os valores numericos na primeira coluna
    mask = dfs.map(lambda x: isinstance(x, (int, float))).iloc[:,0]
    #filtra o dataframe utilizando a máscara
    dfs = dfs[mask]
    #Remove a primeira linha
    dfs = dfs.iloc[1:]
    #Ajusta o cabeçalho
    dfs.columns = ['Ano', 'Mês', 'Índice']
    #Elimina valores expurios na coluna "Mês"
    dfs.dropna(subset=['Mês'], inplace=True)
    #Preenche os valores nan na coluna 'Ano'
    dfs['Ano'] = dfs['Ano'].ffill()
    #Ajusta para string os anos
    dfs['Ano'] = dfs['Ano'].apply(str)

    #Ajusta os meses
    Tabela_mes = pd.DataFrame({'Mês_s': ['JAN','FEV','MAR','ABR','MAI','JUN','JUL','AGO','SET','OUT','NOV','DEZ'], 'Mês_n': ['01','02','03','04','05','06','07','08','09', '10', '11', '12']})
    #Substitui os valores do dataframe conforme a tabela de correspondência
    for i in range(0, len(Tabela_mes)):
        mask = dfs['Mês'] == Tabela_mes.loc[i, 'Mês_s']
        dfs.loc[mask,'Mês_n'] = Tabela_mes.loc[i, 'Mês_n']
        
    #Cria uma nova coluna com a data concatenada conforme metodologia do IBGE (medição do 1º ao 30º dia do mês am análise)
    dfs['Data'] = dfs.apply(lambda x: monta_data(x['Mês_n'], x['Ano']), axis=1)
    #Cria uma coluna com a data no formato de timestamp
    dfs['Data_ts'] = pd.to_datetime(dfs['Data'], format='%d/%m/%Y')
    dfs.reset_index(drop=True, inplace=True)
    return dfs

def ipca_rata(df_ipca):
    #Gera a tabela do índice IPCA pro-rata
    df_ipca_rata = pd.DataFrame([])
    for i in range(0, len(df_ipca)-1):
        idx_ini = i
        idx_fim = i+1
        ni_ini = df_ipca.loc[idx_ini, 'Índice']
        ni_fim = df_ipca.loc[idx_fim, 'Índice']
        data_ini = df_ipca.loc[idx_ini, 'Data_ts']
        data_fim = df_ipca.loc[idx_fim, 'Data_ts']
        datas_inter = gera_datas(data_ini, data_fim)
        #Gera interpolação linear dos valores
        idx_inter = np.linspace(ni_ini, ni_fim, len(datas_inter))
        #Monta o dataframe
        df_aux = pd.DataFrame({'Índice': idx_inter})
        df_aux.index = datas_inter
        #Registra
        df_ipca_rata = pd.concat([df_ipca_rata, df_aux], ignore_index=False)
    df_ipca_rata.drop_duplicates(keep='first', inplace=True)
    return df_ipca_rata

def var_indice(data_ini, data_fim, df_indice):
    #Calcula a variação do índice economico no período especificado
    return (df_indice.loc[data_fim, 'Índice'] / df_indice.loc[data_ini, 'Índice'])

def atualiza_ipca(path_ipca, df_base, db_monet, cols_data, cols_monet):
    #Atualiza monetariamente a base
    #Importa as tabelas de dados do IPCA
    #path_ipca = r'C:/Users/cecil.skaleski/Documents/10. ATR/3_FERRAMENTAS/1_SANEAMENTO/1_FISCALIZAÇÃO/2_AMOSTRAGEM/3_BRR/ipca_202312SerieHist.xls'
    df_ipca = importa_ipca(path_ipca)
    #Monta a tabela com  a variação proporcional do índice (pro-rata)
    df_ipca_rata = ipca_rata(df_ipca)
    
    #Calcula a variação do índice para cada ativo
    df_brr_db = df_base.copy()
    #Define a database monetaria atualizada
    df_brr_db['data_monet_atual'] = pd.to_datetime(db_monet, format='%d/%m/%Y')
    col_ini = cols_data[0]
    col_fim = cols_data[1]
    df_brr_db['var_ipca'] = df_brr_db.apply(lambda x: var_indice(x[col_ini], x[col_fim], df_ipca_rata), axis=1)
    
    #Aplica a variação do índice às colunas selecionadas
    for col in cols_monet:
        df_brr_db[col] = df_brr_db[col] * df_brr_db['var_ipca']
    return  df_brr_db

def TIR(fluxo_caixa):
    #Calcula a TIR de um fluxo de caixa (dataframe)
    tir = npf.irr(fluxo_caixa)
    #Calcula o VPL
    df_vpl = []
    for ano in range(0, len(fluxo_caixa)):
        #Calcula cada elemento do fluxo descontando pela taxa i
        df_vpl.append(fluxo_caixa[ano]/(1 + tir)**ano)
    return df_vpl, tir

def calc_periodos(data_ini, data_fim):
    #Calcula o nº de exercícios entre duas datas e monta a lista de datas de simulação
    #Inicio do exercício: 01/01 (inicio do dia)
    #Fim do exercício: 31/12 (fim do dia)
    
    #Gera a sequência de datas fim do exercício
    ano_ini = data_ini.year
    ano_fim = data_fim.year
    delta_ano = ano_fim - ano_ini
    datas = []
    for x in range(0, delta_ano):
        data = pd.to_datetime(f'31/12/{ano_ini + x}', format='%d/%m/%Y')
        datas.append(data)
    #Adiciona a data final
    datas.append(data_fim)
    return datas
#______________________________________________________________________________________

#______________________________________________________________________________________
#Funções de relatórios
def plota_BRR(df_resumo_brr, n_linhas, rtp, db_monet, db_mov, abs_path, gera_pdf):
    """Plota o valor da BRR bruta e líquida no período de movimentação"""
    csfont = {'fontname':'Calibri'}
    #BRR Bruta e Líquida
    dados_y = df_resumo_brr['BRR bruta']
    ymax = max(dados_y)
    #Trava a escala (para comparações)
    #ymax = 7000000000
    ymin = df_resumo_brr['BRR liquida'].min()
    y_ticks = np.linspace(ymin, ymax, 20)
    y_ticks2 = [millify_rs(y)+'   ' for y in y_ticks]
    deltay = y_ticks[1] - y_ticks[0]
    fig, ax1 = plt.subplots(nrows=1, ncols=1, figsize=(18,8))
    plt.bar(df_resumo_brr['data_imob'], df_resumo_brr['BRR bruta'], color='cornflowerblue', width=0.8, alpha=0.5, label='BRR bruta')
    #plt.plot(df_resumo_brr['data_imob'], df_resumo_brr['BRR bruta'], color='darkblue', alpha=0.8, zorder=10)
    plt.bar(df_resumo_brr['data_imob'], df_resumo_brr['BRR liquida'], color='cornflowerblue', width=0.8, alpha=0.8, label='BRR liquida')
    plt.plot(df_resumo_brr['data_imob'], df_resumo_brr['BRR liquida'], color='darkorange', alpha=0.8, zorder=2)
    plt.xticks(df_resumo_brr['data_imob'], df_resumo_brr['data_imob'].apply(lambda x: x.split('/')[-1]), rotation=80, fontsize=11, color='darkblue', **csfont)
    plt.yticks(y_ticks, y_ticks2, fontsize=11, color='darkblue', **csfont)
    plt.legend(loc='upper right')
    #plt.ylabel('BRR bruta e BRR líquida')
    ax1.spines['bottom'].set_color('white')
    ax1.spines['top'].set_color('white')
    ax1.spines['right'].set_color('white')
    ax1.spines['left'].set_color('white')
    plt.title(f'BRR bruta e BRR líquida (preços de {db_monet})', color='darkblue', fontsize=22, **csfont)
    #Registra
    if gera_pdf == True:
        fname = f"GRAFICO_BRR_{rtp}RTP_DBM-{db_monet.replace('/', '-')}_DBI-{db_mov.strftime('%d-%m-%Y')}_{str(n_linhas)}_itens.pdf"
        folder_path = '6_SAIDA_MOVIMENTA//1_GRAFICOS//'
        path_exp = monta_path(abs_path, folder_path, fname)
        plt.savefig(path_exp, dpi=300, format='pdf')
    plt.show()
    return

def plota_QRR(df_resumo_brr, n_linhas, rtp, db_monet, db_mov, abs_path, gera_pdf):
    """Plota o valor da QRR no período de movimentação"""
    csfont = {'fontname':'Calibri'}
    #QRR
    dados_y = df_resumo_brr['qrr']
    ymax = max(dados_y)
    #Trava a escala (para comparações)
    #ymax = 7000000000
    ymin = min(dados_y)
    y_ticks = np.linspace(ymin, ymax, 20)
    y_ticks2 = [millify_rs(y)+'   ' for y in y_ticks]
    deltay = y_ticks[1] - y_ticks[0]
    fig, ax1 = plt.subplots(nrows=1, ncols=1, figsize=(18,8))
    plt.bar(df_resumo_brr['data_imob'], df_resumo_brr['qrr'], color='cornflowerblue', width=0.8, alpha=0.8, zorder=2)
    plt.plot(df_resumo_brr['data_imob'], df_resumo_brr['qrr'], color='darkblue', alpha=0.3, zorder=1)
    plt.xticks(df_resumo_brr['data_imob'], df_resumo_brr['data_imob'].apply(lambda x: x.split('/')[-1]), rotation=80, fontsize=11, color='darkblue', **csfont)
    plt.yticks(y_ticks, y_ticks2, fontsize=11, color='darkblue', **csfont)
    #plt.legend(loc='best')
    #plt.ylabel('QRR [R$]')
    ax1.spines['bottom'].set_color('white')
    ax1.spines['top'].set_color('white')
    ax1.spines['right'].set_color('white')
    ax1.spines['left'].set_color('white')
    plt.title(f'QRR (preços de {db_monet})', color='darkblue', fontsize=22, **csfont)
    #Registra
    if gera_pdf == True:
        fname = f"GRAFICO_QRR_{rtp}RTP_DBM-{db_monet.replace('/', '-')}_DBI-{db_mov.strftime('%d-%m-%Y')}_{str(n_linhas)}_itens.pdf"
        folder_path = '6_SAIDA_MOVIMENTA//1_GRAFICOS//'
        path_exp = monta_path(abs_path, folder_path, fname)
        plt.savefig(path_exp, dpi=300, format='pdf')
    plt.show()
    return

def plota_TDR(df_resumo_brr, n_linhas, rtp, db_monet, db_mov, abs_path, gera_pdf):
    """Plota o valor da TDR no período de movimentação"""
    csfont = {'fontname':'Calibri'}
    #TDR
    dados_y = df_resumo_brr['tdr_media_anual']
    ymax = max(dados_y)
    #Trava a escala (para comparações)
    #ymax = 7000000000
    ymin = min(dados_y)
    y_ticks = np.linspace(ymin, ymax, 10)
    y_ticks2 = [formats3(y)+'                ' for y in y_ticks]
    deltay = y_ticks[1] - y_ticks[0]
    fig, ax1 = plt.subplots(nrows=1, ncols=1, figsize=(18,4))
    #plt.bar(df_resumo_brr['data_imob'], df_resumo_brr['tdr_media_anual'], color='navy', width=0.8, alpha=0.5, zorder=2)
    plt.plot(df_resumo_brr['data_imob'], df_resumo_brr['tdr_media_anual'], color='navy', alpha=0.4, zorder=1)
    plt.xticks(df_resumo_brr['data_imob'], df_resumo_brr['data_imob'].apply(lambda x: x.split('/')[-1]), rotation=80, fontsize=11, color='navy', **csfont)
    plt.yticks(y_ticks, y_ticks2, fontsize=11, color='navy', **csfont)
    #plt.legend(loc='best')
    plt.ylabel('%')
    ax1.spines['bottom'].set_color('white')
    ax1.spines['top'].set_color('white')
    ax1.spines['right'].set_color('white')
    ax1.spines['left'].set_color('white')
    plt.title('Taxa de depreciação regulatória média anual (TDR)', color='navy', fontsize=22, **csfont)
    #Registra
    if gera_pdf == True:
        fname = f"GRAFICO_TDR_{rtp}RTP_DBM-{db_monet.replace('/', '-')}_DBI-{db_mov.strftime('%d-%m-%Y')}_{str(n_linhas)}_itens.pdf"
        folder_path = '6_SAIDA_MOVIMENTA//1_GRAFICOS//'
        path_exp = monta_path(abs_path, folder_path, fname)
        plt.savefig(path_exp, dpi=300, format='pdf')
    plt.show()
    return
#______________________________________________________________________________________

#______________________________________________________________________________________
#Replica formatação do arquivo excel de saída
def num_para_letra(n):
    letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if n <= 26:
        result = letras[n - 1]
    else:
        quociente = (n - 1) // 26
        resto = (n - 1) % 26
        result =  num_para_letra(quociente) + letras[resto]
    return result

def justificar_tamanho_colunas(ws):
    for coluna in ws.columns:
        max_length = 0
        for cell in coluna[1:]:
            if cell.value:
                if isinstance(cell.value, float):
                    if 0 < cell.value < 1:
                        max_length = max(max_length, 7)
                    else:
                        max_length = max(max_length, len(str(cell.value)))
                else:
                    max_length = max(max_length, len(str(cell.value)))

        largura_justificada = (max_length + 2) * 1.2
        ws.column_dimensions[coluna[0].column_letter].width = largura_justificada
    return

def copiar_estilos_celula(origem, destino):
    destino.font = copy(origem.font)
    destino.fill = copy(origem.fill)
    destino.border = copy(origem.border)
    destino.alignment = copy(origem.alignment)
    destino.protection = copy(origem.protection)
    destino.number_format = origem.number_format
    return

def copia_format(template_path, original_path, output_path):
    template_workbook = openpyxl.load_workbook(template_path)
    original_workbook = openpyxl.load_workbook(original_path)

    template_sheet = template_workbook.active
    original_sheet = original_workbook.active
    
    tpdebug = round(len(list(original_sheet.iter_rows()))/2)
    template_place = list(template_sheet.iter_rows())[1:] * round((len(list(original_sheet.iter_rows()))-1)/round(len(list(template_sheet.iter_rows())[1:])))
    new_template_list = [list(template_sheet.iter_rows())[0]] * round(len(list(original_sheet.iter_rows())[0])/len(list(original_sheet.iter_rows())[0]))
    new_template_list.extend(template_place)

    justificar_tamanho_colunas(original_sheet)

    lenght_x = len(new_template_list[0])
    lenght_y = len(new_template_list)

    for sheet_name in original_workbook.sheetnames:
        ws = original_workbook[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=lenght_y*3, min_col=1, max_col=lenght_x+100):
            for cell in row:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for row_template, row_original in zip(new_template_list, original_sheet.iter_rows()):
        for cell_template, cell_original in zip(row_template, row_original):
            copiar_estilos_celula(cell_template, cell_original)

    original_workbook.save(output_path)
    return
#______________________________________________________________________________________

#______________________________________________________________________________________
#Função principal
def movimenta_BRR(export, gera_pdf, open_folder, abs_path, db_monet, db_mov, anos_sim, path_ref, path_eleg, path_ipca, mensagem_text, output_format_var):
    """Movimenta a BRR conforme orientação técnica (protocolo n° xx.xxx.xx-x)"""
    #Carrega a BRR
    print('')
    print('_____________________________________MOVIMENTA BRR_____________________________________')
    print('Carregando a BRR...')
    mensagem_text.insert(tk.END, '\n\nCarregando a BRR...\n')
    df_brr = importar_arquivo(path_ref)
    #Faz o drop de eventuais valores expurios
    df_brr.dropna(how='all', axis=1, inplace=True)
    df_brr.dropna(how='all', axis=0, inplace=True)
    
    #Importa a lista de alterações de elegibilidade
    mensagem_text.insert(tk.END, '\n\nCarregando a lista de alterações da elegibilidade...\n')
    df_eleg = pd.read_excel(path_eleg)
    #Faz o drop de eventuais valores expurios
    df_eleg.dropna(how='all', axis=1, inplace=True)
    df_eleg.dropna(how='all', axis=0, inplace=True)
    #Converte a data para timestamp
    flag_eleg = False
    if len(df_eleg) > 0:
        df_eleg['data'] = pd.to_datetime(df_eleg['data'])
        flag_eleg = True
    
    #Cria a base de movimentação
    #Dados da base
    #Database monetária atualizada
    df_brr['data_monet_atual'] = pd.to_datetime(db_monet, format='%d/%m/%Y')
    #Database de movimentação atualizada
    df_brr['data_mov_atual'] = pd.to_datetime(db_mov, format='%d/%m/%Y')
    rtp = df_brr['rtp'].apply(int).max()
    
    #Atualiza monetariamente a base
    print('')
    print('____________________________________Atualização monetária____________________________________')
    print(f'Atualizando monetariamente a BRR para a database de {db_monet}...')
    mensagem_text.insert(tk.END, f'\n\n____________________________________Atualização monetária____________________________________\n')
    mensagem_text.insert(tk.END, f'Atualizando monetariamente a BRR para a database de {db_monet}...\n')
    #Cria uma cópia para comparar o resultado
    df_base_bkp = df_brr.copy()
    #Define as colunas que contém as datas iniciais e finais do calculo monetario
    col_ini = 'data_monet'
    col_fim = 'data_monet_atual'
    cols_data = [col_ini, col_fim]   
        
    #Define as colunas que serão atualizadas monetariamente
    cols_monet = ['vrb']
    
    #Atualiza monetarimente
    df_brr_monet = atualiza_ipca(path_ipca, df_brr, db_monet, cols_data, cols_monet)
    
    #Verifica se a atualização foi bem sucedida
    print(f'Verificando se a atualização foi bem sucedida...')
    col_verif = 'var_ipca'
    divers = []
    idx_divers = []
    for col in cols_monet:
       aux = (df_brr_monet[col_verif].apply(lambda x: round(x, 4)) != (df_brr_monet[col] / df_base_bkp[col]).apply(lambda x: round(x, 4)))
       idx_divers += df_brr_monet[aux].index.to_list()
       divers.append(aux.sum())
    #Monta o dataframe
    df_ver_monet = pd.DataFrame({
        'Coluna': cols_monet,
        'Divergência [nº de itens]': divers,
        })
    print('')
    print(df_ver_monet)
    print('')
    print(f'Atualização monetária da BRR concluída com sucesso!')
    print(f"Variação média do índice monetário: {formats3(df_brr_monet['vrb'].sum() / df_brr['vrb'].sum() - 1)}")
    print('')
    print('____________________________________Atualização monetária____________________________________')
    mensagem_text.insert(tk.END, f'\n\n{df_ver_monet}\n\n')
    mensagem_text.insert(tk.END, f'Atualização monetária da BRR concluída com sucesso!\n')
    mensagem_text.insert(tk.END, f"Variação média do índice monetário: {formats3(df_brr_monet['vrb'].sum() / df_brr['vrb'].sum() - 1)}\n")
    mensagem_text.insert(tk.END, f'\n\n____________________________________Atualização monetária____________________________________\n')
    
    #Movimentação da base
    print('')
    print('____________________________________Depreciação da BRR____________________________________')
    mensagem_text.insert(tk.END, f'\n\n____________________________________Depreciação da BRR____________________________________\n')
    #Quadro resumo
    cols_resumo = ['data_imob', 'data_monet_atual', 'Investimento', 'BRR bruta', 'BRR liquida', 'dep_acum_reg', 'saldo_ineleg_bruto', 'saldo_ineleg_liquido']
    #Calcula a depreciação acumulada e o valor regulatório líquido dos ativos
    data_ini = df_brr['data_imob'].min()
    data_fim = pd.to_datetime(db_mov, format='%d/%m/%Y')
    datas = calc_periodos(data_ini, data_fim)
    
    print('')
    print('Depreciando a BRR para as datas selecionadas...')
    mensagem_text.insert(tk.END, f'\n\nDepreciando a BRR para as datas selecionadas...\n')
    results_brr = []
    dfs_brr = []
    df_brr_movs = df_brr_monet.copy()
    for data in datas:
        print(f'\t{data}')
        mensagem_text.insert(tk.END, f'\t{data}\n')
        db_mov = pd.to_datetime(data, format='%d/%m/%Y')
        #Deprecia a base
        if flag_eleg == True:
            df_brr_db, result_brr = calc_brr_imob2(df_brr_movs, db_mov, df_eleg)
        else:
            df_brr_db, result_brr = calc_brr_imob(df_brr_movs, db_mov)
        #Registra
        results_brr.append(result_brr)
        dfs_brr.append(df_brr_db)
    
    #Monta o dataframe
    df_resumo_brr = pd.DataFrame(results_brr)
    df_resumo_brr.columns = cols_resumo
    #Calcula a QRR acumulada em cada período
    qrrs = [df_resumo_brr.loc[0, 'dep_acum_reg']]
    for i in range(1, len(df_resumo_brr)):
        qrr = df_resumo_brr.loc[i, 'dep_acum_reg'] - df_resumo_brr.loc[i-1, 'dep_acum_reg']
        qrrs.append(qrr)
    df_resumo_brr['qrr'] = qrrs
    #Calcula a taxa média de depreciação em cada período
    mask = df_resumo_brr['BRR bruta'] > 0
    df_resumo_brr['tdr_media_anual'] = 0.0
    df_resumo_brr.loc[mask, 'tdr_media_anual'] = (df_resumo_brr.loc[mask, 'qrr'] / df_resumo_brr.loc[mask, 'BRR bruta'])
    #Calcula os totais
    qrr_total = df_resumo_brr['qrr'].sum()
    
    #Monta fluxo de caixa de verificação
    #Avalia os investimentos realizados no período (crescimento da BRR bruta)
    df_fc = df_resumo_brr.copy()
    df_fc['Investimento'] = -1 * df_resumo_brr['Investimento']
     
    #Repete a ultima linha e adiciona o valor residual, se houver
    last_date = df_fc['data_imob'].tail(1).iloc[0]                                            
    df_fc = pd.concat([df_fc, df_fc.tail(1)], axis=0, ignore_index=True)
    #Ajusta a data do ultimo ano do fluxo
    df_fc.loc[df_fc.index[-1], 'data_imob'] = (pd.to_datetime(last_date, dayfirst=True) + pd.DateOffset(years=1)).strftime('%d/%m/%Y') 
    df_fc['Amortização'] = df_fc['qrr']
    #Ajusta a amortização incluindo eventual saldo não amortizado
    df_fc.loc[df_fc.index[-1], 'Amortização'] = df_fc.loc[df_fc.index[-1], 'BRR liquida']
    #Ajusta o investimento no último ano (desconsidera)
    df_fc.loc[df_fc.index[-1], 'Investimento'] = 0
    
    #Calcula os juros
    tx_juros = 0.1182768 # Taxa de juros 
    n_alg = len(str(tx_juros*100).split('.')[-1])
    juros = [0] + (df_resumo_brr['BRR liquida']*tx_juros).to_list()
    df_fc['Juros'] = juros
    #Inclui o saldo devedor
    df_fc['Saldo'] = df_fc['BRR liquida']
    #Calcula o fluxo de pagamentos
    df_fc['Fluxo'] = df_fc['Investimento'] + df_fc['Amortização'] + df_fc['Juros']
    
    #Adiciona as informações do fluxo de caixa ao quadro resumo
    df_resumo_brr['Saldo'] = df_fc['Saldo']
    
    #Ajusta o nº de períodos do fluxo de caixa
    aux_corte = df_fc[df_fc['Fluxo'] == 0]
    if len(aux_corte) > 0:
        idx_corte = aux_corte.head(1).index[0]
        df_fc = df_fc.head(idx_corte)
    
    #Calcula a TIR do fluxo gerado
    fluxo_caixa = df_fc['Fluxo'].to_list()
    df_vpl, tir = TIR(fluxo_caixa)
    # Avalia o erro relativo na estimativa da TIR
    erro_tir = ((tir - tx_juros) / tx_juros)  # Diferença já em valores percentuais
    tol_tir = 0.001 / 100 # Tolerancia já em valores percentuais (1 = 1%)
    
    #Cria uma cópia para exibição do quadro resumo na tela
    df_resumo_brr_fmt = df_resumo_brr.copy()
    #Formata os valores monetários
    cols_format = ['Investimento', 'BRR bruta', 'BRR liquida', 'dep_acum_reg', 'saldo_ineleg_bruto',  'saldo_ineleg_liquido', 'qrr', 'Saldo']
    for col in cols_format:
        df_resumo_brr_fmt[col] = df_resumo_brr_fmt[col].apply(formats2)
    print('')
    print(df_resumo_brr_fmt)
    print('')
    print(f"Base depreciada com sucesso! Valor da depreciação acumulada até {data}: {df_resumo_brr_fmt['dep_acum_reg'].tail(1).iloc[0]}")
    print(f"QRR total paga até {data}: {formats2(qrr_total)}")
    #Calcula o investimento total realizado no período
    tot_invest = df_resumo_brr['Investimento'].sum()
    print(f'Total de investimentos imobilizados até {data}: {formats2(tot_invest)}')
    mensagem_text.insert(tk.END, f'\n\n{df_resumo_brr_fmt}\n\n')
    mensagem_text.insert(tk.END, f"Base depreciada com sucesso! Valor da depreciação acumulada até {data}: {df_resumo_brr_fmt['dep_acum_reg'].tail(1).iloc[0]}\n")
    mensagem_text.insert(tk.END, f"QRR total paga até {data}: {formats2(qrr_total)}\n")
    #Calcula o investimento total realizado no período
    tot_invest = df_resumo_brr['Investimento'].sum()
    mensagem_text.insert(tk.END, f'Total de investimentos imobilizados até {data}: {formats2(tot_invest)}\n')
    
    #Seleciona as colunas para compor a tabela resumo
    cols_mov = [
    'iu',
    'rtp',
    'municipio',
    'servico',
    'conta_contabil',
    'descricao', 
    'qtde',
    'data_imob',
    'data_monet',
    'vrb',
    'vrl',
    'tdr_percent_ano',
    ]
    df_brr_mov = df_brr_db[cols_mov]
    
    #Calcula a diferença entre as bases bruta e líquida
    dif_brr = df_brr_mov['vrb'].sum() - df_brr_mov['vrl'].sum()
    print(f'Diferença BRR bruta-líquida em {data}: {formats2(dif_brr)}')
    print('')
    print('____________________________________Depreciação da BRR____________________________________')
    mensagem_text.insert(tk.END, f'Diferença BRR bruta-líquida em {data}: {formats2(dif_brr)}\n\n')
    mensagem_text.insert(tk.END, f'\n\n____________________________________Depreciação da BRR____________________________________\n')
    
    #Apresenta as informações do fluxo de caixa
    print(' ')
    print(f'Taxa de juros utilizada: {str(tx_juros*100)}%')
    print(f'TIR estimada: {round(tir*100, n_alg)}%')
    mensagem_text.insert(tk.END, ' ')
    mensagem_text.insert(tk.END, f'Taxa de juros utilizada: {str(tx_juros*100)}%\n')
    mensagem_text.insert(tk.END, f'TIR estimada: {round(tir*100, n_alg)}%\n')
    if abs(erro_tir) < tol_tir:
        print(f'Erro na estimativa da TIR: {formats3(erro_tir)}, inferior a {tol_tir*100}%')
        mensagem_text.insert(tk.END, f'Erro na estimativa da TIR: {formats3(erro_tir)}, inferior a {tol_tir*100}%\n')
    else:
        print(f'Erro na estimativa da TIR: {formats3(erro_tir)}, superior ao limite de {tol_tir*100}%')
        mensagem_text.insert(tk.END, f'Erro na estimativa da TIR: {formats3(erro_tir)}, superior ao limite de {tol_tir*100}%\n')
    #_______________________________________
    
    #Apresenta os gráficos na tela
    #BRR bruta e líquida
    n_linhas = len(df_brr_mov)
    plota_BRR(df_resumo_brr, n_linhas, rtp, db_monet, db_mov, abs_path, gera_pdf)
    #QRR
    plota_QRR(df_resumo_brr, n_linhas, rtp, db_monet, db_mov, abs_path, gera_pdf)
    #TDR
    plota_TDR(df_resumo_brr, n_linhas, rtp, db_monet, db_mov, abs_path, gera_pdf)
    
    #Apresenta uma lista dos ativos que não depreciaram 100% (conferência de valores)
    df_nao_deprec = df_brr_db[(df_brr_db['vrl'].apply(round) != 0) & (df_brr_db['elegibilidade'] != 'Não elegível')]
    #Separa ativos não amortizáveis dos amortizáveis
    df_nao_amort = df_nao_deprec[df_nao_deprec['taxa_deprec_anos'] == 0]
    df_amort = df_nao_deprec[df_nao_deprec['taxa_deprec_anos'] > 0]
    pd.set_option('display.max_rows', 300)
    print('')
    print(f"Ativos não amortizáveis: {len(df_nao_amort)} ({formats2(df_nao_amort['vrl'].sum())})")
    mensagem_text.insert(tk.END, f"\n\nAtivos não amortizáveis: {len(df_nao_amort)} ({formats2(df_nao_amort['vrl'].sum())})\n")
    if len(df_nao_amort) > 0:
        print('')
        print(agrupa2(df_nao_amort, [6, 5, 12], 27, 1, 1))
        print('')
        print('')
        mensagem_text.insert(tk.END, f"\n{agrupa2(df_nao_amort, [6, 5, 12], 27, 1, 1)}\n\n")
    
    print(f"Ativos com saldo a amortizar: {len(df_amort)} ({formats2(df_amort['vrl'].sum())})")
    print('')
    mensagem_text.insert(tk.END, f"Ativos com saldo a amortizar: {len(df_amort)} ({formats2(df_amort['vrl'].sum())})\n")
    if len(df_amort) > 0:
        print(agrupa2(df_amort, [6, 5, 12], 27, 1, 1))
        mensagem_text.insert(tk.END, f"\n{agrupa2(df_amort, [6, 5, 12], 27, 1, 1)}\n")
    
    #Exporta resultados
    #Exporta o resumo da movimentação da BRR em formato de planilha excel
    export = True
    if export == True:
        print(f'Exportando dados em formato .xlsx...')
        mensagem_text.insert(tk.END, f'\nExportando dados em formato .xlsx...\n')
        fname = f"RESUMO_BRR_{rtp}RTP_DBM-{db_monet.replace('/', '-')}_DBI-{db_mov.strftime('%d-%m-%Y')}_{str(n_linhas)}_itens.xlsx"
        folder_path = '6_SAIDA_MOVIMENTA//'
        path_exp = monta_path(abs_path, folder_path, fname)
        exportar_arquivo(df_resumo_brr, path_exp, output_format_var)
        print(f'    Arquivo exportado com sucesso!')
        print(f'    {path_exp}')
        mensagem_text.insert(tk.END, f'    Arquivo exportado com sucesso!\n')
        mensagem_text.insert(tk.END, f'    {path_exp}\n')
        #Replica a formatação do template
        print('Replicando a formatação do arquivo de modelo...')
        mensagem_text.insert(tk.END, "Replicando a formatação do arquivo de modelo...\n")
        mensagem_text.see('end')
        mensagem_text.update()
        folder_path = '5_ENTRADA_MOVIMENTA/1_FORMATOS//'
        fname = f'Template_resumo_brr.{output_format_var}'
        template_path = monta_path(abs_path, folder_path, fname)
        copia_format(template_path, path_exp, path_exp)
        print('    Formatação replicada com sucesso!')
        mensagem_text.insert(tk.END, "    Formatação replicada com sucesso!\n")
        mensagem_text.see('end')
        mensagem_text.update()
        
        #Exporta a BRR final em formato de planilha excel
        n_linhas = len(df_brr_mov)
        print(f'Exportando dados em formato .xlsx...')
        mensagem_text.insert(tk.END, f'Exportando dados em formato .xlsx...\n')
        fname = f"BRR_{rtp}RTP_DBM-{db_monet.replace('/', '-')}_DBI-{db_mov.strftime('%d-%m-%Y')}_{str(n_linhas)}_itens.xlsx"
        folder_path = '6_SAIDA_MOVIMENTA//'
        folder_exp = folder_path.split('/')[0]
        path_exp = monta_path(abs_path, folder_path, fname)
        exportar_arquivo(df_brr_mov, path_exp, output_format_var)
        print(f'    Arquivo exportado com sucesso!')
        print(f'    {path_exp}')
        mensagem_text.insert(tk.END, f'    Arquivo exportado com sucesso!\n')
        mensagem_text.insert(tk.END, f'    {path_exp}\n')
        
        #Abre a pasta com os arquivos gerados
        if open_folder == True:
            res = subprocess.Popen(fr'explorer "{folder_exp}"')
    #_______________________________________
    print('_____________________________________MOVIMENTA BRR_____________________________________')
    return
#______________________________________________________________________________________

#______________________________________________________________________________________
#Funções GUI
def buscar_arquivo(entry_arquivo, path_var):
    abs_path = os.path.dirname(__file__)
    folder_path = '5_ENTRADA_MOVIMENTA'
    dir_ini = os.path.join(abs_path, folder_path)
    arquivo = filedialog.askopenfilename(initialdir=dir_ini, title='Selecione o arquivo')
    entry_arquivo.delete(0, tk.END)
    entry_arquivo.insert(0, arquivo)
    entry_arquivo.xview_moveto(1)
    path_var.set(arquivo)

def movimentar_brr(db_monet_var, db_mov_var, anos_sim_var, path_ref_var, path_eleg_var, path_ipca_var, mensagem_text, output_format_var):
    #Captura o conteúdo dos elementos de texto
    db_monet = db_monet_var.get()
    db_mov = db_mov_var.get()
    #anos_sim = anos_sim_var.get()
    path_ref = path_ref_var.get()
    path_eleg = path_eleg_var.get()
    path_ipca = path_ipca_var.get()
    #______________________________________EXECUÇÃO________________________________________
    #Seleciona os arquivos de entrada
    abs_path = os.path.dirname(__file__)
    export = True
    gera_pdf = True
    open_folder = True
    anos_sim = anos_sim_var
    movimenta_BRR(export, gera_pdf, open_folder, abs_path, db_monet, db_mov, anos_sim, path_ref, path_eleg, path_ipca, mensagem_text, output_format_var)
    #______________________________________EXECUÇÃO________________________________________
    #Feedback na caixa de mensagens
    mensagem_text.config(state=tk.NORMAL)
    mensagem_text.insert(tk.END, "Base de dados movimentada com sucesso!\n")
    mensagem_text.config(state=tk.DISABLED)

def make_frame(frame):
    #Captura o conteudo das caixas de texto com o caminho dos arquivos selecionados
    path_ref_var = tk.StringVar()
    path_eleg_var = tk.StringVar()
    path_ipca_var = tk.StringVar()
    anos_sim_var = 76
    output_format_var = tk.StringVar(value="xlsx")  # Valor padrão
    
    # Arquivo de BRR
    label_arquivo = tk.Label(frame, text="Arquivo de BRR:", fg="blue")
    label_arquivo.grid(row=0, column=0, padx=10, pady=(10,0), sticky="w")

    entry_arquivo1 = tk.Entry(frame, width=65, bg='lightgrey')
    entry_arquivo1.grid(row=0, column=1, pady=(10,0), sticky="w", padx=10)

    btn_buscar = tk.Button(frame, text="Buscar", bg="navy", fg="white", command=lambda: buscar_arquivo(entry_arquivo1, path_ref_var))
    btn_buscar.grid(row=0, column=2, padx=0, pady=(10,0), sticky='w')

    # Tabela Elegibilidade
    label_tabela_de_para = tk.Label(frame, text="Tabela elegibilidade:", fg="blue")
    label_tabela_de_para.grid(row=1, column=0, padx=10, sticky="w")

    entry_arquivo2 = tk.Entry(frame, width=65, bg='lightgrey')
    entry_arquivo2.grid(row=1, column=1, sticky="w", padx=10)

    btn_buscar = tk.Button(frame, text="Buscar", bg="navy", fg="white", command=lambda: buscar_arquivo(entry_arquivo2, path_eleg_var))
    btn_buscar.grid(row=1, column=2, padx=0, sticky='w')

    # Tabela IPCA
    label_plano_contas = tk.Label(frame, text="Tabela IPCA:", fg="blue")
    label_plano_contas.grid(row=2, column=0, padx=10, sticky="w")

    entry_arquivo3 = tk.Entry(frame, width=65, bg='lightgrey')
    entry_arquivo3.grid(row=2, column=1, sticky="w", padx=10)

    btn_buscar = tk.Button(frame, text="Buscar", bg="navy", fg="white", command=lambda: buscar_arquivo(entry_arquivo3, path_ipca_var))
    btn_buscar.grid(row=2, column=2, padx=0, sticky='w')

    # Database monetária
    label_data_monetaria = tk.Label(frame, text="Database monetária:", fg="blue")
    label_data_monetaria.grid(row=3, column=0, padx=(10, 0), sticky="w")

    entry_data_monetaria = tk.Entry(frame, width=20, bg='lightgrey')
    entry_data_monetaria.grid(row=3, column=1, padx=10, sticky="w")

    # Database Movimentação
    label_data_mov = tk.Label(frame, text="Database movimentação:", fg="blue")
    label_data_mov.grid(row=4, column=0, padx=(10, 0), sticky="w")

    entry_data_mov = tk.Entry(frame, width=20, bg='lightgrey')
    entry_data_mov.grid(row=4, column=1, padx=10, sticky="w")

    #Dropdown para seleção do formato de saída
    label_formato = tk.Label(frame, text="Formato de Saída:", fg="blue")
    label_formato.grid(row=5, column=0, padx=10, pady=10, sticky="w")

    formatos = ["xlsx", "csv", "hdf", "feather"]
    dropdown_formatos = tk.OptionMenu(frame, output_format_var, *formatos)
    dropdown_formatos.config(width=10, bg="lightgrey")
    dropdown_formatos.grid(row=5, column=1, padx=10, sticky="w")
    
    # Botão Movimentar BRR
    btn_converter_base_dados = tk.Button(frame, text="Movimentar BRR", bg="navy", fg="white", width=20, height=2, command=lambda: movimentar_brr(entry_data_monetaria, entry_data_mov, anos_sim_var, path_ref_var, path_eleg_var, path_ipca_var, mensagem_text, output_format_var.get()))
    btn_converter_base_dados.grid(row=6, column=0, columnspan=3, pady=10)

    # Display
    mensagem_label = tk.Label(frame, text="Mensagens:", fg="blue")
    mensagem_label.grid(row=7, column=0, pady=(0, 0), padx=10, sticky="w")

    mensagem_text = tk.Text(frame, height=3, width=80, bg="white", bd=1, relief="solid")
    mensagem_text.grid(row=8, column=0, columnspan=3, padx=10, pady=5, sticky="ew")
    mensagem_text.config(state=tk.DISABLED)

    scrollbar = tk.Scrollbar(frame, command=mensagem_text.yview)
    scrollbar.grid(row=8, column=3, rowspan=3, sticky='nse')
    mensagem_text['yscrollcommand'] = scrollbar.set

def init_frame():
    root = tk.Tk()
    root.title("Movimenta BRR")

    largura_janela = 700
    altura_janela = 310

    largura_tela = root.winfo_screenwidth()
    altura_tela = root.winfo_screenheight()

    posx = largura_tela // 2 - largura_janela // 2
    posy = altura_tela // 2 - altura_janela // 2

    root.geometry(f"{largura_janela}x{altura_janela}+{posx}+{posy}")

    frame = tk.Frame(root)
    frame.pack(pady=20)

    make_frame(frame)

    root.mainloop()
