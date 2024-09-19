# -*- coding: utf-8 -*-
"""
Created on Thu May 16 10:16:32 2024

@author: cecil.skaleski, est.angelo

Converte a base de dados da BRR confome disposto na Nota Técnica n° XX/2024.        
"""

import math
import pandas as pd
from pandas.api.types import is_string_dtype
from pandas.api.types import is_numeric_dtype
import numpy as np
import matplotlib.pyplot as plt
import os
from dateutil.relativedelta import relativedelta
import datetime
import matplotlib.dates as md
import subprocess

import openpyxl
from openpyxl.styles import PatternFill
from copy import copy

import tkinter as tk
from tkinter import filedialog

np.set_printoptions(linewidth=np.inf)
pd.set_option('display.max_columns', 20)
pd.set_option('display.max_rows', 100)
pd.set_option('display.width', 100)
pd.set_option("display.precision", 6)
# Don't wrap repr(DataFrame) across additional lines
pd.set_option("display.expand_frame_repr", False)
pd.options.display.float_format = '{:.2f}'.format
pd.set_option("display.date_dayfirst", True)
pd.set_option('display.max_colwidth', 55)

#______________________________________________________________________________________
#Funções acessórias
def formats2(x):
    return ('R${:,.2f}'.format(x)).replace(",", "~").replace(".", ",").replace("~", ".")

def formats3(x):
    return "{:.2f}%".format(x*100)

def millify(n):
    millnames = ['',' Mil',' Milhões',' Bilhões',' Trilhões']
    n = float(n)
    millidx = max(0,min(len(millnames)-1,
                        int(math.floor(0 if n == 0 else math.log10(abs(n))/3))))
    return '{:.2f}{}'.format(n / 10**(3 * millidx), millnames[millidx]).replace('.', ',')

def millify_rs(n):
    millnames = ['',' Mil',' Milhões',' Bilhões',' Trilhões']
    n = float(n)
    millidx = max(0,min(len(millnames)-1,
                        int(math.floor(0 if n == 0 else math.log10(abs(n))/3))))
    return 'R${:.2f}{}'.format(n / 10**(3 * millidx), millnames[millidx]).replace('.', ',')

def aplica_strip(x):
    #Aplica a função strip
    if isinstance(x, str):
        x = x.strip()
    return x

def remove_espacos(database):
    #Remove os espaços em branco do início e do final de todos os campos do dataframe que contenham strings
    #Retorna o relatório das strings alteradas, por coluna e quantidade
    #Guarda o nome de cada coluna do dataframe
    colunas = database.columns.to_list()
    aux = database.copy()
    #Analisa o tipo de dado em cada coluna
    for i in colunas:
        if is_string_dtype(aux.loc[:, i]):
            aux.loc[:, i] = aux.loc[:, i].apply(aplica_strip)
    return aux

def monta_path(abs_path, folder_path, fname):
    #Retorna o caminho absoluto para o arquivo
    return os.path.join(abs_path, folder_path+fname)

def escolhe_arq(titulo, filetypes, dir_ini):
    root = tk.Tk()
    root.wm_attributes('-topmost', 1)
    root.withdraw()
    file_path = filedialog.askopenfilename(parent=root, initialdir=dir_ini, title=titulo, filetypes=filetypes)
    #showinfo(title='Arquivo selecionado', message=filename))
    root.destroy()
    return file_path

def exportar_arquivo(df, output_path, output_format):
    if output_format == "xlsx":
        with pd.ExcelWriter(output_path) as writer:
            df.to_excel(writer, index=False)
    elif output_format == "csv":
        df.to_csv(output_path, index=False)
    elif output_format == "hdf":
        df.to_hdf(output_path, key='df', mode='w')
    elif output_format == "feather":
        df.to_feather(output_path)
    else:
        raise ValueError(f"Formato de saída '{output_format}' não suportado!")
#______________________________________________________________________________________

#______________________________________________________________________________________
#Funções de processamento de dados
def importa_plano_contas(path):
    #Importa o plano de contas em formato excel
    #Importa arquivo excel (alterar a engine e comando de importação para outros formatos)
    dfs = pd.read_excel(path, sheet_name=0)
    #Remove as linhas com valores NaN
    dfs = dfs.dropna()
    #Remove os espaços em branco
    dfs = remove_espacos(dfs)
    return dfs

def insere_plano_contas2(database, indice_conta_contabil, df_plano_contas, nome_plano):
    #Insere o plano de contas no final do dataframe
    aux = database.copy()
    #Cria a nova coluna no final do dataframe copiando a coluna com os codigos contabeis
    nome_col = aux.columns[indice_conta_contabil]
    aux[nome_plano] = None
    #Varre o dataframe na coluna indicada, alimentando a coluna do plano de contas fazendo a correspondencia com a tabela do plano de contas
    for i in range(0, len(df_plano_contas)):
        #Procura na coluna de referencia o codigo
        mask = (aux.loc[:, nome_col] == df_plano_contas.iloc[i, 0])
        #Substitui pelo valor de correspondencia
        aux.loc[mask, nome_plano] = df_plano_contas.iloc[i, 1]
        #print(f'Codigo {df_plano_contas.iloc[i, 0]} substituido por {df_plano_contas.iloc[i, 1]}')
    return aux
       
def iu(tipo, plaqueta, complemento):
    #Monta o código identificador único do ativo na base de referência da Sanepar
    return f"{str(tipo).split('.')[0]}-{str(plaqueta).split('.')[0]}-{str(complemento).split('.')[0]}"

def tabela_resumo(df, contas, col_qtde, col_custo, col_conta, col_mun):
    #Calcula a tabela resumo
    qtdes = []
    muns = []
    custos = []
    itens = []
    for conta in contas:
        df_aux = df[df[col_conta] == conta]
        item = len(df_aux)
        if item > 0:
            qtde = df_aux[col_qtde].sum()
            nmun = len(df_aux[col_mun].unique())
            custo = df_aux[col_custo].sum()
        else:
            qtde = 0
            nmun = 0
            custo = 0
        #Registra
        qtdes.append(qtde)
        muns.append(nmun)
        custos.append(custo)
        itens.append(item)
    #Monta o dataframe
    df_resumo = pd.DataFrame({
        'Conta contábil': contas,
        'Linhas': itens,
        'Qtde de bens': qtdes,
        'N municípios': muns,
        'Custo contábil': custos,
        })
    #Ordena por maior custo, linhas, qtde e municipios
    df_resumo = df_resumo.sort_values(['Custo contábil', 'Linhas', 'Qtde de bens', 'N municípios'], ascending=[False, False, False, False]).reset_index(drop=True)
    #Calcula o impacto percentual e cumulativo
    df_resumo['%'] = df_resumo['Custo contábil'] / df_resumo['Custo contábil'].sum()
    df_resumo['% acum'] = df_resumo['%'].cumsum()
    return df_resumo

def verifica_reqs(df, col_names):
    #Verifica a consistência das colunas indicadas (iu, taxa de depreciação e qtde)
    col_iu = col_names[0]
    col_deprec = col_names[1]
    col_qtde = col_names[2]
    flag = False
    err_msg = []
    #Verifica se a coluna existe
    if col_iu not in df.columns:
        err_msg.append('Coluna com identificador único (iu) não encontrada!')
        flag = True
    else:
        #Verifica se a coluna é consistente
        mask = (df[col_iu].isnull()) | (df[col_iu].apply(str).str.strip() == '')
        df_iu = df[mask]
        if len(df_iu) > 0:
            err_msg.append('Coluna identificador único (iu) inconsistente! (itens nulos ou vazios)')
            flag = True
        #Verifica se há itens com iu replicado
        df_aux = df.groupby(col_iu).size().reset_index(name='qtde')
        df_rep = df_aux[df_aux['qtde'] > 1]
        if len(df_rep) > 0:
            err_msg.append('Identificadores únicos (iu) replicados nos dados processados!')
            flag = True 
        #Verifica a composição do código identificador único de ativo (IU)
        df_comp = df.copy()
        df_comp['iu_verif'] = df_comp.apply(lambda x: iu(x['tipo'], x['plaqueta'], x['complemento']), axis=1)
        #Compara
        df_comp_res = df_comp[df_comp['iu_verif'] != df_comp[col_iu]]
        if len(df_comp_res) > 0:
            #Itens com código inadequado
            print('')
            print('Inconsistências detectadas na composição do identificador único: ')
            print(df_comp_res.loc[:, ['tipo', 'plaqueta', 'complemento', col_iu, 'iu_verif']])
            err_msg.append('Inconsistências detectadas na composição do identificador único!')
            flag = True
    #Itens sem taxa de depreciação
    #Verifica se a coluna existe
    if col_deprec not in df.columns:
        err_msg.append('Coluna taxa de depreciação não encontrada!')
        flag = True
    else:
        #Verifica se a coluna é consistente
        mask = (df[col_deprec].isnull()) | (df[col_deprec].apply(str).str.strip() == '')
        df_iu = df[mask]
        if len(df_iu) > 0:
            err_msg.append('Coluna taxa de depreciação inconsistente! (itens nulos ou vazios)')
            flag = True
    #Itens sem quantidade definida
    #Verifica se a coluna existe
    if col_qtde not in df.columns:
        err_msg.append('Coluna quantidade não encontrada!')
        flag = True
    else:
        #Verifica se a coluna é consistente
        mask = (df[col_qtde].isnull()) | (df[col_qtde].apply(str).str.strip() == '')
        df_iu = df[mask]
        if len(df_iu) > 0:
            err_msg.append('Coluna quantidade inconsistente! (itens nulos ou vazios)')
            flag = True
    return flag, err_msg
#______________________________________________________________________________________

#______________________________________________________________________________________
#Replica formatação do arquivo excel de saída
def num_para_letra(n):
    letras = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if n <= 26:
        result = letras[n - 1]
    else:
        quociente = (n - 1) // 26
        resto = (n - 1) % 26
        result =  num_para_letra(quociente) + letras[resto]
    return result

def justificar_tamanho_colunas(ws):
    for coluna in ws.columns:
        max_length = 0
        for cell in coluna[1:]:
            if cell.value:
                if isinstance(cell.value, float):
                    if 0 < cell.value < 1:
                        max_length = max(max_length, 7)
                    else:
                        max_length = max(max_length, len(str(cell.value)))
                else:
                    max_length = max(max_length, len(str(cell.value)))

        largura_justificada = (max_length + 2) * 1.2
        ws.column_dimensions[coluna[0].column_letter].width = largura_justificada
    return

def copiar_estilos_celula(origem, destino):
    destino.font = copy(origem.font)
    destino.fill = copy(origem.fill)
    destino.border = copy(origem.border)
    destino.alignment = copy(origem.alignment)
    destino.protection = copy(origem.protection)
    destino.number_format = origem.number_format
    return

def copia_format(template_path, original_path, output_path):
    template_workbook = openpyxl.load_workbook(template_path)
    original_workbook = openpyxl.load_workbook(original_path)

    template_sheet = template_workbook.active
    original_sheet = original_workbook.active
    
    tpdebug = round(len(list(original_sheet.iter_rows()))/2)
    template_place = list(template_sheet.iter_rows())[1:] * round((len(list(original_sheet.iter_rows()))-1)/round(len(list(template_sheet.iter_rows())[1:])))
    new_template_list = [list(template_sheet.iter_rows())[0]] * round(len(list(original_sheet.iter_rows())[0])/len(list(original_sheet.iter_rows())[0]))
    new_template_list.extend(template_place)

    justificar_tamanho_colunas(original_sheet)

    lenght_x = len(new_template_list[0])
    lenght_y = len(new_template_list)

    for sheet_name in original_workbook.sheetnames:
        ws = original_workbook[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=lenght_y*3, min_col=1, max_col=lenght_x+100):
            for cell in row:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    for row_template, row_original in zip(new_template_list, original_sheet.iter_rows()):
        for cell_template, cell_original in zip(row_template, row_original):
            copiar_estilos_celula(cell_template, cell_original)

    original_workbook.save(output_path)
    return
#______________________________________________________________________________________

#______________________________________________________________________________________
#Função principal
def converte_BRR(export, open_folder, abs_path, path_ref, path_dp, path_contas, mensagem_text, output_format_var):
    """Converte a base de dados em BRR parcial conforme orientação técnica (protocolo n° xx.xxx.xx-x)"""
    #Habilita o feedback na caixa de mensagens
    mensagem_text.config(state=tk.NORMAL)
    #_______________________________________
    #Carrega a base de dados
    print('')
    print('_____________________________________CONVERTE BRR_____________________________________')
    print('Carregando base de dados...')
    mensagem_text.insert(tk.END, "\n\nCarregando base de dados...\n")
    mensagem_text.see('end')
    mensagem_text.update()
    #folder_path = '1_ENTRADA_CONVERTE'
    #dir_ini = os.path.join(abs_path, folder_path)
    #tipo_arqs = [('XLSX', '.xlsx'), ('XLS', '.xls'), ('CSV', '.csv'), ('JSON', '.json'), ('ALL', '.*')]
    #path_ref = escolhe_arq('Selecione o arquivo da base de dados', tipo_arqs, dir_ini)
    df_ref = pd.read_excel(path_ref, sheet_name='BRR Incremental')
    df_ref.dropna(how='all', axis=1, inplace=True)
    df_ref.dropna(how='all', axis=0, inplace=True)
    
    #Importa a tabela "Depara" da base de dados
    print('Carregando tabela de conversão de atributos...')
    mensagem_text.insert(tk.END, "Carregando tabela de conversão de atributos...\n")
    mensagem_text.see('end')
    mensagem_text.update()
    #path_dp = escolhe_arq('Selecione o arquivo da tabela "De Para"', [('XLSX', '.xlsx'), ('CSV', '.csv'), ('JSON', '.json')], dir_ini)
    df_dp = pd.read_excel(path_dp)
    df_dp.dropna(how='all', axis=1, inplace=True)
    df_dp.dropna(how='all', axis=0, inplace=True)
    #_______________________________________
    
    #_______________________________________
    #Procedimentos de tratamento da base de dados
    #Insere o plano de contas da Sanepar
    #tipo_arqs = [('XLSX', '.xlsx'), ('XLS', '.xls'), ('CSV', '.csv'), ('JSON', '.json'), ('ALL', '.*')]
    #path_contas = escolhe_arq('Selecione o arquivo com o plano de contas', tipo_arqs, dir_ini)
    print('Carregando o plano de contas da Sanepar...')
    mensagem_text.insert(tk.END, "Carregando o plano de contas da Sanepar...\n")
    mensagem_text.see('end')
    mensagem_text.update()
    df_contas = importa_plano_contas(path_contas)
    df_contas.dropna(how='all', axis=1, inplace=True)
    df_contas.dropna(how='all', axis=0, inplace=True)
    print('Inserindo o plano de contas da Sanepar...')
    mensagem_text.insert(tk.END, "Inserindo o plano de contas da Sanepar...\n")
    mensagem_text.see('end')
    mensagem_text.update()
    n_conta = df_ref.columns.to_list().index('ANALISE') #indice da coluna "ANALISE"
    df_ref = insere_plano_contas2(df_ref, n_conta, df_contas, 'CONTA CONTABIL (DESCRICAO)')
    print('    Plano de contas inserido com sucesso!')
    mensagem_text.insert(tk.END, "    Plano de contas inserido com sucesso!\n")
    mensagem_text.see('end')
    mensagem_text.update()
    #Explicita o ano de imobilização
    df_ref['ano_imob'] = df_ref['DT CONTABIL'].apply(lambda x: str(x.year))
    
    #Seleciona as colunas indicadas na tabela "Depara"
    cols_origem = df_dp['col_origem'].to_list()
    cols_brr = df_dp['col_brr'].to_list()
    df_base = df_ref.loc[:, cols_origem].copy()
    df_base.columns = cols_brr
    #_______________________________________
    
    #_______________________________________
    #Verificação de requisitos mínimos
    print('')
    print('Verificando requisitos mínimos de consistência de dados...')
    mensagem_text.insert(tk.END, "\n\nVerificando requisitos mínimos de consistência de dados...\n")
    mensagem_text.see('end')
    mensagem_text.update()
    flag, err_msg = verifica_reqs(df_base, ['iu', 'taxa_deprec_anos', 'qtde'])
    flag_cont = 'S'
    if flag == True:
        print('')
        for msgm in err_msg:
            print(msgm)
        flag_cont = (input('Requisitos de consistência não cumpridos. Continuar mesmo assim? (S ou N) ')).upper().strip()
        mensagem_text.insert(tk.END, "\nRequisitos de consistência não cumpridos. Continuar mesmo assim?\n")
        mensagem_text.see('end')
        mensagem_text.update()
    else:
        print('    Requisitos verificados!')
        mensagem_text.insert(tk.END, "    Requisitos verificados!\n")
        mensagem_text.see('end')
        mensagem_text.update()
    if flag_cont == 'S':   
        #_______________________________________
        #Cria tabela resumo por conta contábil
        #Avalia a quantidade de itens e os valores totais por conta contábil
        col_qtde = 'qtde'
        col_custo = 'custo_contabil'
        col_conta = 'conta_contabil'
        col_mun = 'municipio'
        
        #Monta a tabela de resumo
        contas = df_base[col_conta].unique()
        df_res_ref = tabela_resumo(df_base, contas, col_qtde, col_custo, col_conta, col_mun)
        #Adiciona os totais
        df_res_ref.loc['TOTAL', :] = df_res_ref.sum(axis=0)
        df_res_ref.loc['TOTAL', 'Conta contábil'] = 'TOTAL'
        df_res_ref.loc['TOTAL', 'N municípios'] = len(df_base[col_mun].unique())
        df_res_ref.loc['TOTAL', '% acum'] = 1
        #Formata valores unitários
        df_res_ref['Linhas'] = df_res_ref['Linhas'].apply(int)
        df_res_ref['Qtde de bens'] = df_res_ref['Qtde de bens'].apply(int)
        df_res_ref['N municípios'] = df_res_ref['N municípios'].apply(int)
        
        #Apresenta o resultado
        print('')
        print('Tabela resumo da base de referência')
        df_res_ref_f = df_res_ref.copy()
        #Formata valores monetários
        df_res_ref_f['Custo contábil'] = df_res_ref_f['Custo contábil'].apply(formats2)
        #Formata valores percentuais
        df_res_ref_f['%'] = df_res_ref_f['%'].apply(formats3)
        df_res_ref_f['% acum'] = df_res_ref_f['% acum'].apply(formats3)
        print(df_res_ref_f)
        #_______________________________________
        
        #_______________________________________
        #Exporta resultados
        rtp = df_base['rtp'].apply(int).max()
        n_linhas = len(df_base)
        data_hj = datetime.datetime.today().strftime('%d-%m-%Y_%Hh%Mmin%Ss')
        #Exporta o resumo da BRR em formato de planilha excel
        if export == True:
            print('')
            print(f'Exportando o resumo da BRR parcial em formato .{output_format_var}...')
            mensagem_text.insert(tk.END, "\n\nExportando o resumo da BRR parcial em formato .{output_format_var}...\n")
            mensagem_text.update()
            fname = f"RESUMO_BRR_PARCIAL_{rtp}RTP_{str(n_linhas)}_itens_{data_hj}.{output_format_var}"
            folder_path = '2_SAIDA_CONVERTE//'
            path_exp = monta_path(abs_path, folder_path, fname)
            exportar_arquivo(df_res_ref, path_exp, output_format_var)
            print(f'    Arquivo exportado com sucesso!')
            print(f'    {path_exp}')
            #Replica a formatação do template
            print('Replicando a formatação do arquivo de modelo...')
            mensagem_text.insert(tk.END, "Replicando a formatação do arquivo de modelo...\n")
            mensagem_text.see('end')
            mensagem_text.update()
            folder_path = '1_ENTRADA_CONVERTE/1_FORMATOS//'
            fname = f'Template_resumo_brr_parcial.{output_format_var}'
            template_path = monta_path(abs_path, folder_path, fname)
            copia_format(template_path, path_exp, path_exp)
            print('    Formatação replicada com sucesso!')
            mensagem_text.insert(tk.END, "    Formatação replicada com sucesso!\n")
            mensagem_text.see('end')
            mensagem_text.update()
            
        #Exporta a BRR parcial em formato excel
            print('')
            print(f'Exportando dados da BRR parcial em formato .{output_format_var}...')
            mensagem_text.insert(tk.END, f"\n\nExportando dados da BRR parcial em formato .{output_format_var}...\n")
            mensagem_text.see('end')
            mensagem_text.update()
            fname = f"BRR_PARCIAL_{rtp}RTP_{str(n_linhas)}_itens_{data_hj}.{output_format_var}"
            folder_path = '2_SAIDA_CONVERTE//'
            folder_exp = folder_path.split('/')[0]
            path_exp = monta_path(abs_path, folder_path, fname)
            exportar_arquivo(df_base, path_exp, output_format_var)
            print(f'    Arquivo exportado com sucesso!')
            print(f'    {path_exp}')
            mensagem_text.insert(tk.END, f"    Arquivo exportado com sucesso! {path_exp}\n")
            mensagem_text.see('end')
            mensagem_text.update()
            
            #Replica a formatação do template
            #print('Replicando a formatação do arquivo de modelo...')
            #folder_path = '1_ENTRADA_CONVERTE/1_FORMATOS//'
            #fname = 'Template_brr_parcial.xlsx'
            #template_path = monta_path(abs_path, folder_path, fname)
            #copia_format(template_path, path_exp, path_exp) #Desabilitado por tempo de execução excessivo
            #print('    Formatação replicada com sucesso!')
        #Abre a pasta com os arquivos gerados
            if open_folder == True:
                res = subprocess.Popen(fr'explorer "{folder_exp}"')
        #_______________________________________
    #_______________________________________
    #Desabilita o feedback no campo de mensagens
    mensagem_text.config(state=tk.DISABLED)
    print('_____________________________________CONVERTE BRR_____________________________________')
    return
#______________________________________________________________________________________

#______________________________________________________________________________________
#Funções GUI
def buscar_arquivo(entry_arquivo, path_var):
    abs_path = os.path.dirname(__file__)
    folder_path = '1_ENTRADA_CONVERTE'
    dir_ini = os.path.join(abs_path, folder_path)
    arquivo = filedialog.askopenfilename(initialdir=dir_ini, title='Selecione o arquivo')
    entry_arquivo.delete(0, tk.END)
    entry_arquivo.insert(0, arquivo)
    entry_arquivo.xview_moveto(1)
    path_var.set(arquivo)

def printm(texto, mensagem_text):
    print(texto)
    mensagem_text.insert(tk.END, texto)

def converter_base_dados(path_ref_var, path_dp_var, path_contas_var, mensagem_text, output_format_var):
    #Captura o conteúdo dos elementos de texto
    path_ref = path_ref_var.get()
    path_dp = path_dp_var.get()
    path_contas = path_contas_var.get()
    #______________________________________EXECUÇÃO________________________________________
    #Seleciona os arquivos de entrada
    abs_path = os.path.dirname(__file__)
    export = True
    open_folder = True
    converte_BRR(export, open_folder, abs_path, path_ref, path_dp, path_contas, mensagem_text, output_format_var)
    #______________________________________EXECUÇÃO________________________________________
    #Feedback na caixa de mensagens
    mensagem_text.config(state=tk.NORMAL)
    mensagem_text.insert(tk.END, "Base de dados convertida com sucesso!\n")
    mensagem_text.config(state=tk.DISABLED)

def make_frame(frame):
    #Captura o conteudo das caixas de texto com o caminho dos arquivos selecionados
    path_ref_var = tk.StringVar()
    path_dp_var = tk.StringVar()
    path_contas_var = tk.StringVar()
    output_format_var = tk.StringVar(value="xlsx")  # Valor padrão

    # Arquivo de Dados
    label_arquivo = tk.Label(frame, text="Arquivo de Dados:", fg="blue")
    label_arquivo.grid(row=0, column=0, padx=10, pady=(10,0), sticky="w")

    entry_arquivo1 = tk.Entry(frame, width=70, bg='lightgrey', textvariable=path_ref_var)
    entry_arquivo1.grid(row=0, column=1, pady=(10,0))

    btn_buscar1 = tk.Button(frame, text="Buscar", bg="navy", fg="white", command=lambda: buscar_arquivo(entry_arquivo1, path_ref_var))
    btn_buscar1.grid(row=0, column=2, padx=10, pady=(10,0), sticky="w")

    # Tabela De-Para
    label_tabela_de_para = tk.Label(frame, text="Tabela De-Para:", fg="blue")
    label_tabela_de_para.grid(row=1, column=0, padx=10, sticky="w")

    entry_arquivo2 = tk.Entry(frame, width=70, bg='lightgrey', textvariable=path_dp_var)
    entry_arquivo2.grid(row=1, column=1)

    btn_buscar2 = tk.Button(frame, text="Buscar", bg="navy", fg="white", command=lambda: buscar_arquivo(entry_arquivo2, path_dp_var))
    btn_buscar2.grid(row=1, column=2, padx=10, sticky="w")

    # Plano de Contas
    label_plano_contas = tk.Label(frame, text="Plano de Contas:", fg="blue")
    label_plano_contas.grid(row=2, column=0, padx=10, sticky="w")

    entry_arquivo3 = tk.Entry(frame, width=70, bg='lightgrey', textvariable=path_contas_var)
    entry_arquivo3.grid(row=2, column=1)

    btn_buscar3 = tk.Button(frame, text="Buscar", bg="navy", fg="white", command=lambda: buscar_arquivo(entry_arquivo3, path_contas_var))
    btn_buscar3.grid(row=2, column=2, padx=10, sticky="w")

    #Dropdown para seleção do formato de saída
    label_formato = tk.Label(frame, text="Formato de Saída:", fg="blue")
    label_formato.grid(row=3, column=0, padx=10, pady=10, sticky="w")

    formatos = ["xlsx", "csv", "hdf", "feather"]
    dropdown_formatos = tk.OptionMenu(frame, output_format_var, *formatos)
    dropdown_formatos.config(width=10, bg="lightgrey")
    dropdown_formatos.grid(row=3, column=1, padx=7, sticky="w")

    # Botão Converter Base de Dados
    btn_converter_base_dados = tk.Button(frame, text="Converter Base de Dados", bg="navy", fg="white", width=20, height=2, command=lambda: converter_base_dados(path_ref_var, path_dp_var, path_contas_var, mensagem_text, output_format_var.get()))
    btn_converter_base_dados.grid(row=4, column=1, columnspan=3, pady=10, sticky="e")

    # Display
    mensagem_label = tk.Label(frame, text="Mensagens:", fg="blue")
    mensagem_label.grid(row=6, column=0, pady=(40, 0), padx=10, sticky="w")

    mensagem_text = tk.Text(frame, height=3, width=80, bg="white", bd=1, relief="solid")
    mensagem_text.grid(row=7, column=0, columnspan=3, padx=10, pady=5, sticky="ew")
    mensagem_text.config(state=tk.DISABLED)
    
    scrollbar = tk.Scrollbar(frame, command=mensagem_text.yview)
    scrollbar.grid(row=7, column=3, rowspan=3, sticky='nse')
    mensagem_text['yscrollcommand'] = scrollbar.set

def init_frame():
    root = tk.Tk()
    root.title("Converte BRR")

    largura_janela = 700
    altura_janela = 280

    largura_tela = root.winfo_screenwidth()
    altura_tela = root.winfo_screenheight()

    posx = largura_tela // 2 - largura_janela // 2
    posy = altura_tela // 2 - altura_janela // 2

    root.geometry(f"{largura_janela}x{altura_janela}+{posx}+{posy}")

    frame = tk.Frame(root)
    frame.pack(pady=20)

    make_frame(frame)

    root.mainloop()
